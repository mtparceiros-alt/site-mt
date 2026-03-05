"""
MT Parceiros - Gerador de Planilha Mestre v2.0
Gera MT_Parceiros_Painel_AAAA-MM.xlsx com formatação profissional,
validações, dashboard com fórmulas dinâmicas e aba de configurações.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime


# ─────────────────────────────────────────────
# PALETA DE CORES MT PARCEIROS
# ─────────────────────────────────────────────
COR_LARANJA       = "F35525"   # Primária — cabeçalhos principais
COR_LARANJA_CLARO = "FDEBD0"   # Fundo alternado quente
COR_CINZA_ESCURO  = "2C2C2C"   # Texto sobre fundo laranja
COR_CINZA_CLARO   = "F5F5F5"   # Fundo alternado frio
COR_BRANCO        = "FFFFFF"
COR_VERDE         = "1E8449"   # Status positivo
COR_VERDE_CLARO   = "D5F5E3"
COR_VERMELHO      = "C0392B"   # Status negativo
COR_VERMELHO_CLARO= "FADBD8"
COR_AZUL          = "1A5276"   # Dashboard
COR_AZUL_CLARO    = "D6EAF8"
COR_AMARELO       = "F9E79F"   # Linha de exemplo / destaque

FONTE_PADRAO = "Arial"
LINHAS_DADOS = 500  # Linhas pré-formatadas para entrada de dados


# ─────────────────────────────────────────────
# HELPERS DE ESTILO
# ─────────────────────────────────────────────
def cabecalho(texto_cor=COR_BRANCO, fundo_cor=COR_LARANJA, negrito=True, tamanho=10, centralizar=True):
    alinhamento = Alignment(horizontal="center" if centralizar else "left",
                            vertical="center", wrap_text=True)
    return {
        "font":      Font(name=FONTE_PADRAO, bold=negrito, color=texto_cor, size=tamanho),
        "fill":      PatternFill("solid", fgColor=fundo_cor),
        "alignment": alinhamento,
        "border":    borda_fina(),
    }

def celula_normal(cor_fundo=COR_BRANCO):
    return {
        "font":      Font(name=FONTE_PADRAO, size=10),
        "fill":      PatternFill("solid", fgColor=cor_fundo),
        "alignment": Alignment(vertical="center"),
        "border":    borda_fina(),
    }

def borda_fina():
    lado = Side(style="thin", color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def borda_media():
    lado = Side(style="medium", color=COR_LARANJA)
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def aplicar_estilo(cell, estilo: dict):
    for attr, val in estilo.items():
        setattr(cell, attr, val)

def formatar_cabecalho_linha(ws, linha, colunas, cor_fundo=COR_LARANJA, cor_texto=COR_BRANCO):
    for col in range(1, colunas + 1):
        cell = ws.cell(row=linha, column=col)
        aplicar_estilo(cell, cabecalho(cor_texto, cor_fundo))

def freeze_and_filter(ws, cell_ref, filter_range):
    ws.freeze_panes = cell_ref
    ws.auto_filter.ref = filter_range


# ─────────────────────────────────────────────
# ABA: CONFIGURAÇÕES (oculta — fonte dos dropdowns)
# ─────────────────────────────────────────────
def criar_aba_configuracoes(wb):
    ws = wb.create_sheet("CONFIGURACOES")
    ws.sheet_state = "hidden"

    listas = {
        "A": ("STATUS_IMOVEL",    ["Lançamento", "Em obras", "Pronto", "Suspenso"]),
        "B": ("STATUS_VENDA",     ["Novo", "Em contato", "Proposta enviada", "Negociando", "Vendido", "Perdido"]),
        "C": ("CANAL",            ["WhatsApp", "Telefone", "Site", "Indicação", "Instagram", "Visita direta"]),
        "D": ("RESULTADO",        ["Positivo", "Neutro", "Negativo", "Sem resposta"]),
        "E": ("FAIXA_MCMV",       ["Faixa 1", "Faixa 2", "Faixa 3", "Faixa 4", "SBPE/Mercado"]),
    }

    for col_letter, (titulo, opcoes) in listas.items():
        ws[f"{col_letter}1"] = titulo
        aplicar_estilo(ws[f"{col_letter}1"], cabecalho(fundo_cor=COR_AZUL))
        for i, opcao in enumerate(opcoes, start=2):
            ws[f"{col_letter}{i}"] = opcao

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    return ws


# ─────────────────────────────────────────────
# ABA: IMOVEIS
# ─────────────────────────────────────────────
def criar_aba_imoveis(wb):
    ws = wb.create_sheet("IMOVEIS")

    # Título
    ws.merge_cells("A1:I1")
    ws["A1"] = "🏠  CADASTRO DE IMÓVEIS — MT PARCEIROS"
    aplicar_estilo(ws["A1"], cabecalho(tamanho=12, fundo_cor=COR_LARANJA))

    # Cabeçalhos
    colunas = [
        ("A", "ID_IMOVEL",   8),
        ("B", "NOME",        30),
        ("C", "STATUS",      16),
        ("D", "BAIRRO",      20),
        ("E", "ENDEREÇO",    35),
        ("F", "LATITUDE",    12),
        ("G", "LONGITUDE",   12),
        ("H", "PRECO_BASE",  16),
        ("I", "URL_IMAGEM",  35),
    ]
    for col, titulo, largura in colunas:
        cell = ws[f"{col}2"]
        cell.value = titulo
        aplicar_estilo(cell, cabecalho(fundo_cor=COR_CINZA_ESCURO))
        ws.column_dimensions[col].width = largura

    # Linhas de dados com fundo alternado + ID automático
    for row in range(3, LINHAS_DADOS + 3):
        cor = COR_BRANCO if row % 2 == 0 else COR_CINZA_CLARO
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            aplicar_estilo(cell, celula_normal(cor))

        # ID automático: só preenche se NOME estiver preenchido
        ws.cell(row=row, column=1).value = f'=IF(B{row}<>"",ROW()-2,"")'
        ws.cell(row=row, column=1).font = Font(name=FONTE_PADRAO, size=10, color="0000FF")

        # Formato de moeda na coluna H (PRECO_BASE)
        ws.cell(row=row, column=8).number_format = 'R$ #,##0'

        # Formato de número nas coords
        ws.cell(row=row, column=6).number_format = '0.0000000'
        ws.cell(row=row, column=7).number_format = '0.0000000'

    # Validação de STATUS (dropdown)
    dv_status = DataValidation(
        type="list",
        formula1="CONFIGURACOES!$A$2:$A$5",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Status inválido",
        error="Selecione um status da lista."
    )
    ws.add_data_validation(dv_status)
    dv_status.sqref = f"C3:C{LINHAS_DADOS + 2}"

    # Freeze e filtro
    freeze_and_filter(ws, "A3", f"A2:I2")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 36

    return ws


# ─────────────────────────────────────────────
# ABA: LEADS
# ─────────────────────────────────────────────
def criar_aba_leads(wb):
    ws = wb.create_sheet("LEADS")

    ws.merge_cells("A1:K1")
    ws["A1"] = "👤  GESTÃO DE LEADS — MT PARCEIROS"
    aplicar_estilo(ws["A1"], cabecalho(tamanho=12, fundo_cor=COR_LARANJA))

    colunas = [
        ("A", "ID_LEAD",              8),
        ("B", "DATA_ENTRADA",         14),
        ("C", "NOME_CLIENTE",         28),
        ("D", "WHATSAPP",             16),
        ("E", "ID_IMOVEL_INTERESSE",  10),
        ("F", "FAIXA_MCMV",          14),
        ("G", "POTENCIAL_COMPRA",     18),
        ("H", "STATUS_VENDA",         18),
        ("I", "ULTIMA_INTERACAO",     16),
        ("J", "DATA_RETORNO",         14),
        ("K", "OBSERVACOES",          35),
    ]
    for col, titulo, largura in colunas:
        cell = ws[f"{col}2"]
        cell.value = titulo
        aplicar_estilo(cell, cabecalho(fundo_cor=COR_CINZA_ESCURO))
        ws.column_dimensions[col].width = largura

    for row in range(3, LINHAS_DADOS + 3):
        cor = COR_BRANCO if row % 2 == 0 else COR_LARANJA_CLARO

        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            aplicar_estilo(cell, celula_normal(cor))

        # ID automático
        ws.cell(row=row, column=1).value = f'=IF(C{row}<>"",ROW()-2,"")'
        ws.cell(row=row, column=1).font = Font(name=FONTE_PADRAO, size=10, color="0000FF")

        # Formatos de data
        ws.cell(row=row, column=2).number_format  = 'DD/MM/AAAA'
        ws.cell(row=row, column=9).number_format  = 'DD/MM/AAAA'
        ws.cell(row=row, column=10).number_format = 'DD/MM/AAAA'

        # Formato moeda
        ws.cell(row=row, column=7).number_format = 'R$ #,##0'

    # Dropdown STATUS_VENDA
    dv_status = DataValidation(
        type="list",
        formula1="CONFIGURACOES!$B$2:$B$7",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Status inválido",
        error="Selecione um status da lista."
    )
    ws.add_data_validation(dv_status)
    dv_status.sqref = f"H3:H{LINHAS_DADOS + 2}"

    # Dropdown FAIXA_MCMV
    dv_faixa = DataValidation(
        type="list",
        formula1="CONFIGURACOES!$E$2:$E$6",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Faixa inválida",
        error="Selecione uma faixa da lista."
    )
    ws.add_data_validation(dv_faixa)
    dv_faixa.sqref = f"F3:F{LINHAS_DADOS + 2}"

    # Formatação condicional: linha verde se Vendido, vermelha se Perdido
    range_leads = f"A3:K{LINHAS_DADOS + 2}"
    ws.conditional_formatting.add(range_leads, FormulaRule(
        formula=[f'$H3="Vendido"'],
        fill=PatternFill("solid", fgColor=COR_VERDE_CLARO),
        font=Font(name=FONTE_PADRAO, color=COR_VERDE)
    ))
    ws.conditional_formatting.add(range_leads, FormulaRule(
        formula=[f'$H3="Perdido"'],
        fill=PatternFill("solid", fgColor=COR_VERMELHO_CLARO),
        font=Font(name=FONTE_PADRAO, color=COR_VERMELHO)
    ))

    freeze_and_filter(ws, "A3", f"A2:K2")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 36

    return ws


# ─────────────────────────────────────────────
# ABA: INTERAÇÕES
# ─────────────────────────────────────────────
def criar_aba_interacoes(wb):
    ws = wb.create_sheet("INTERACOES")

    ws.merge_cells("A1:G1")
    ws["A1"] = "💬  HISTÓRICO DE INTERAÇÕES — MT PARCEIROS"
    aplicar_estilo(ws["A1"], cabecalho(tamanho=12, fundo_cor=COR_LARANJA))

    colunas = [
        ("A", "ID_INTERACAO", 12),
        ("B", "ID_LEAD",      10),
        ("C", "DATA_HORA",    18),
        ("D", "CANAL",        18),
        ("E", "NOTAS",        45),
        ("F", "RESULTADO",    16),
        ("G", "PROXIMO_PASSO",35),
    ]
    for col, titulo, largura in colunas:
        cell = ws[f"{col}2"]
        cell.value = titulo
        aplicar_estilo(cell, cabecalho(fundo_cor=COR_CINZA_ESCURO))
        ws.column_dimensions[col].width = largura

    for row in range(3, LINHAS_DADOS + 3):
        cor = COR_BRANCO if row % 2 == 0 else COR_AZUL_CLARO
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            aplicar_estilo(cell, celula_normal(cor))

        ws.cell(row=row, column=1).value = f'=IF(B{row}<>"",ROW()-2,"")'
        ws.cell(row=row, column=1).font = Font(name=FONTE_PADRAO, size=10, color="0000FF")
        ws.cell(row=row, column=3).number_format = 'DD/MM/AAAA HH:MM'

        # Quebra de texto para NOTAS
        ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical="top")

    # Dropdown CANAL
    dv_canal = DataValidation(
        type="list",
        formula1="CONFIGURACOES!$C$2:$C$7",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Canal inválido",
        error="Selecione um canal da lista."
    )
    ws.add_data_validation(dv_canal)
    dv_canal.sqref = f"D3:D{LINHAS_DADOS + 2}"

    # Dropdown RESULTADO
    dv_resultado = DataValidation(
        type="list",
        formula1="CONFIGURACOES!$D$2:$D$5",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Resultado inválido",
        error="Selecione um resultado da lista."
    )
    ws.add_data_validation(dv_resultado)
    dv_resultado.sqref = f"F3:F{LINHAS_DADOS + 2}"

    freeze_and_filter(ws, "A3", f"A2:G2")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 36

    return ws


# ─────────────────────────────────────────────
# ABA: DASHBOARD
# ─────────────────────────────────────────────
def criar_aba_dashboard(wb):
    ws = wb.create_sheet("DASHBOARD")

    # ── Título principal ──
    ws.merge_cells("B2:H2")
    ws["B2"] = "📊  PAINEL DE CONTROLE — MT PARCEIROS"
    ws["B2"].font      = Font(name=FONTE_PADRAO, bold=True, size=14, color=COR_BRANCO)
    ws["B2"].fill      = PatternFill("solid", fgColor=COR_LARANJA)
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 36

    # ── Gerado em ──
    ws.merge_cells("B3:H3")
    ws["B3"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["B3"].font      = Font(name=FONTE_PADRAO, italic=True, size=9, color="888888")
    ws["B3"].alignment = Alignment(horizontal="right")

    # ── Larguras ──
    ws.column_dimensions["A"].width = 3
    for col in ["B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 18

    def card_titulo(ws, row, col, texto, cor_fundo=COR_AZUL):
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
        cell = ws.cell(row=row, column=col, value=texto)
        cell.font      = Font(name=FONTE_PADRAO, bold=True, size=10, color=COR_BRANCO)
        cell.fill      = PatternFill("solid", fgColor=cor_fundo)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 22

    def card_valor(ws, row, col, formula, formato="geral", cor_fundo=COR_AZUL_CLARO):
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
        cell = ws.cell(row=row, column=col, value=formula)
        cell.font      = Font(name=FONTE_PADRAO, bold=True, size=18, color=COR_AZUL)
        cell.fill      = PatternFill("solid", fgColor=cor_fundo)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if formato == "moeda":
            cell.number_format = 'R$ #,##0'
        elif formato == "pct":
            cell.number_format = '0.0%'
        ws.row_dimensions[row].height = 38
        borda = borda_media()
        for c in range(col, col+2):
            ws.cell(row=row, column=c).border = borda

    def label_valor(ws, row, col, label, formula, formato="geral"):
        ws.cell(row=row, column=col).value = label
        ws.cell(row=row, column=col).font = Font(name=FONTE_PADRAO, size=9, color="555555")
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="right", vertical="center")

        cell_val = ws.cell(row=row, column=col+1, value=formula)
        cell_val.font = Font(name=FONTE_PADRAO, bold=True, size=10, color=COR_CINZA_ESCURO)
        cell_val.alignment = Alignment(horizontal="left", vertical="center")
        if formato == "moeda":
            cell_val.number_format = 'R$ #,##0'
        elif formato == "pct":
            cell_val.number_format = '0.0%'
        ws.row_dimensions[row].height = 20

    # ── BLOCO 1: Leads ──
    card_titulo(ws, 5,  2, "TOTAL DE LEADS",         COR_AZUL)
    card_valor( ws, 6,  2, f'=COUNTA(LEADS!C3:C{LINHAS_DADOS+2})')
    card_titulo(ws, 5,  4, "LEADS ESTE MÊS",          COR_AZUL)
    card_valor( ws, 6,  4,
        f'=COUNTIFS(LEADS!B3:B{LINHAS_DADOS+2},">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),'
        f'LEADS!B3:B{LINHAS_DADOS+2},"<="&TODAY())')
    card_titulo(ws, 5,  6, "VENDAS REALIZADAS",       COR_VERDE)
    card_valor( ws, 6,  6,
        f'=COUNTIF(LEADS!H3:H{LINHAS_DADOS+2},"Vendido")',
        cor_fundo=COR_VERDE_CLARO)

    # ── BLOCO 2: Conversão e Potencial ──
    card_titulo(ws, 9,  2, "TAXA DE CONVERSÃO",       COR_LARANJA)
    card_valor( ws, 10, 2,
        f'=IFERROR(COUNTIF(LEADS!H3:H{LINHAS_DADOS+2},"Vendido")'
        f'/COUNTA(LEADS!C3:C{LINHAS_DADOS+2}),0)',
        formato="pct", cor_fundo=COR_LARANJA_CLARO)
    card_titulo(ws, 9,  4, "POTENCIAL MÉDIO",         COR_LARANJA)
    card_valor( ws, 10, 4,
        f'=IFERROR(AVERAGEIF(LEADS!G3:G{LINHAS_DADOS+2},">"&0),0)',
        formato="moeda", cor_fundo=COR_LARANJA_CLARO)
    card_titulo(ws, 9,  6, "POTENCIAL TOTAL (CARTEIRA)", COR_LARANJA)
    card_valor( ws, 10, 6,
        f'=SUMIF(LEADS!H3:H{LINHAS_DADOS+2},"<>Perdido",LEADS!G3:G{LINHAS_DADOS+2})',
        formato="moeda", cor_fundo=COR_LARANJA_CLARO)

    # ── BLOCO 3: Funil de vendas ──
    ws.merge_cells("B13:H13")
    ws["B13"] = "FUNIL DE VENDAS"
    ws["B13"].font      = Font(name=FONTE_PADRAO, bold=True, size=11, color=COR_BRANCO)
    ws["B13"].fill      = PatternFill("solid", fgColor=COR_CINZA_ESCURO)
    ws["B13"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[13].height = 24

    status_funil = [
        ("Novo",              COR_AZUL_CLARO),
        ("Em contato",        "D6EAF8"),
        ("Proposta enviada",  COR_AMARELO),
        ("Negociando",        "FAD7A0"),
        ("Vendido",           COR_VERDE_CLARO),
        ("Perdido",           COR_VERMELHO_CLARO),
    ]
    for i, (status, cor) in enumerate(status_funil):
        row = 14 + i
        label_valor(ws, row, 2, status,
            f'=COUNTIF(LEADS!H3:H{LINHAS_DADOS+2},"{status}")')
        for c in range(2, 8):
            ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=cor)
        ws.row_dimensions[row].height = 20

    # ── BLOCO 4: Imóveis por status ──
    ws.merge_cells("B22:H22")
    ws["B22"] = "IMÓVEIS POR STATUS"
    ws["B22"].font      = Font(name=FONTE_PADRAO, bold=True, size=11, color=COR_BRANCO)
    ws["B22"].fill      = PatternFill("solid", fgColor=COR_CINZA_ESCURO)
    ws["B22"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[22].height = 24

    status_imoveis = ["Lançamento", "Em obras", "Pronto", "Suspenso"]
    for i, status in enumerate(status_imoveis):
        row = 23 + i
        label_valor(ws, row, 2, status,
            f'=COUNTIF(IMOVEIS!C3:C{LINHAS_DADOS+2},"{status}")')
        ws.row_dimensions[row].height = 20

    # ── BLOCO 5: Canais de origem ──
    ws.merge_cells("B29:H29")
    ws["B29"] = "LEADS POR CANAL DE ORIGEM"
    ws["B29"].font      = Font(name=FONTE_PADRAO, bold=True, size=11, color=COR_BRANCO)
    ws["B29"].fill      = PatternFill("solid", fgColor=COR_CINZA_ESCURO)
    ws["B29"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[29].height = 24

    # Canal vem de INTERACOES, relacionado via ID_LEAD
    canais = ["WhatsApp", "Telefone", "Site", "Indicação", "Instagram", "Visita direta"]
    for i, canal in enumerate(canais):
        row = 30 + i
        label_valor(ws, row, 2, canal,
            f'=COUNTIF(INTERACOES!D3:D{LINHAS_DADOS+2},"{canal}")')
        ws.row_dimensions[row].height = 20

    # ── Nota de rodapé ──
    ws.merge_cells("B38:H38")
    ws["B38"] = "⚠️  Os valores são calculados automaticamente. Não edite esta aba manualmente."
    ws["B38"].font      = Font(name=FONTE_PADRAO, italic=True, size=8, color="999999")
    ws["B38"].alignment = Alignment(horizontal="center")

    ws.sheet_view.showGridLines = False

    return ws


# ─────────────────────────────────────────────
# MONTAGEM FINAL
# ─────────────────────────────────────────────
def gerar_planilha():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove aba padrão vazia

    # Ordem das abas
    criar_aba_dashboard(wb)
    criar_aba_leads(wb)
    criar_aba_imoveis(wb)
    criar_aba_interacoes(wb)
    criar_aba_configuracoes(wb)

    # Aba ativa ao abrir = Dashboard
    wb.active = wb["DASHBOARD"]

    # Nome com data para evitar sobrescrever versões anteriores
    data_atual = datetime.now().strftime("%Y-%m")
    nome_arquivo = f"MT_Parceiros_Painel_{data_atual}.xlsx"

    wb.save(nome_arquivo)
    print(f"✅  Planilha gerada: {nome_arquivo}")
    return nome_arquivo


if __name__ == "__main__":
    gerar_planilha()
