# ============================================================
#  sync_cms.py — MT Parceiros CMS
#  Lê o Empreendimentos.xlsx e gera empreendimentos.js
#  com TODOS os campos (imagem, diferenciais, lazer, mapa).
# ============================================================
import json, time, sys, os, re
import urllib.request, urllib.parse

try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl")
    import openpyxl

# ── Configurações ────────────────────────────────────────────
ARQUIVO_XLSX   = "Empreendimentos.xlsx"
ARQUIVO_JS     = "empreendimentos.js"
ARQUIVO_CACHE  = "coords_cache.json"
PASTA_IMAGENS  = "imag"

# Correções Manuais de Coordenadas (GPS validado)
FIXES = {
    "Rua Laguna, 440":                     {"lat": -23.63959, "lng": -46.71574},
    "Avenida Cristo Rei, 34":              {"lat": -23.48938, "lng": -46.72078},
    "Professor Miguel Franchini Neto, 186": {"lat": -23.44355, "lng": -46.72530},
    "Rua Professor Clemente Pastore, 14":  {"lat": -23.51383, "lng": -46.69922},
    "Rua Joao Alfredo, 431":               {"lat": -23.65701, "lng": -46.70044},
    "Avenida Condessa Elisabeth de Robiano, 5748": {"lat": -23.51590, "lng": -46.55550}
}

# Limites de São Paulo
LAT_MIN, LAT_MAX = -23.85, -23.35
LNG_MIN, LNG_MAX = -46.95, -46.35

# Índices das colunas no Excel
COL_NOME    = 0   # A: Empreendimento
COL_BAIRRO  = 1   # B: Região / Bairro
COL_END     = 2   # C: Endereço
COL_AREA    = 3   # D: Área
COL_QUARTOS = 4   # E: Quartos
COL_DIFER   = 5   # F: Opções / Diferenciais
COL_LAZER   = 6   # G: Destaques de Lazer / Facilidades
COL_PRECO   = 7   # H: A partir de
COL_ENTREGA = 8   # I: Entrega
COL_IMAGEM  = 9   # J: Imagem

# ── Carregar Variáveis de Ambiente (.env) ────────────────────
def carregar_env():
    env = {}
    if os.path.exists(".env"):
        with open(".env", "r", encoding="utf-8") as f:
            for line in f:
                if "=" in line and not line.startswith("#"):
                    k, v = line.strip().split("=", 1)
                    env[k.strip()] = v.strip()
    return env

ENV = carregar_env()
GOOGLE_KEY = ENV.get("GOOGLE_MAPS_API_KEY")

# ── Funções ──────────────────────────────────────────────────
def carregar_cache():
    if os.path.exists(ARQUIVO_CACHE):
        try:
            with open(ARQUIVO_CACHE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return {}
    return {}

def salvar_cache(cache):
    with open(ARQUIVO_CACHE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)

def consultar_viacep(cep):
    """Consulta o ViaCEP para validar o endereço antes da geocodificação."""
    if not cep: return None
    cep_limpo = re.sub(r"\D", "", str(cep))
    if len(cep_limpo) != 8: return None
    
    try:
        url = f"https://viacep.com.br/ws/{cep_limpo}/json/"
        with urllib.request.urlopen(url, timeout=5) as resp:
            data = json.loads(resp.read().decode())
            if "erro" not in data:
                return {
                    "logradouro": data.get("logradouro", ""),
                    "bairro": data.get("bairro", ""),
                    "cidade": data.get("localidade", ""),
                    "uf": data.get("uf", "")
                }
    except: pass
    return None

def geocodificar_google(endereco, bairro):
    if not GOOGLE_KEY:
        return None, None, None
        
    addr_clean = f"{endereco}, {bairro}, Sao Paulo, SP, Brazil"
    url = f"https://maps.googleapis.com/maps/api/geocode/json?address={urllib.parse.quote(addr_clean)}&key={GOOGLE_KEY}"
    
    try:
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode())
            if data["status"] == "OK":
                res = data["results"][0]
                lat = res["geometry"]["location"]["lat"]
                lng = res["geometry"]["location"]["lng"]
                return lat, lng, "Google"
    except:
        pass
    return None, None, None

def geocodificar_nominatim(endereco, bairro, cep_info=None):
    headers = { 'User-Agent': 'MTParceirosGeocoder/6.0' }
    addr_clean = endereco.replace("\n", " ").strip()

    # Checar correções fixas
    for partial_match, coords in FIXES.items():
        if partial_match.lower() in addr_clean.lower():
            return coords["lat"], coords["lng"], "Fixa"

    # Preparar consultas variadas
    consultas = []
    
    # 1. Se temos info do CEP, usar a rua oficial confirmada
    if cep_info and cep_info['logradouro']:
        # Extrair o número do endereço original (ex: "Rua X, 123" -> "123")
        numero = re.search(r'(\d+)', addr_clean)
        num_str = numero.group(1) if numero else ""
        consultas.append(f"{cep_info['logradouro']} {num_str}, {cep_info['cidade']}, {cep_info['uf']}, Brazil")

    # 2. Consultas padrão
    consultas.append(f"{addr_clean}, Sao Paulo, SP, Brazil")
    consultas.append(f"{addr_clean}, {bairro}, Sao Paulo, Brazil")
    
    for q in consultas:
        try:
            # Busca estruturada via q=
            q_encoded = urllib.parse.quote(q)
            url = f"https://nominatim.openstreetmap.org/search?q={q_encoded}&format=json&limit=3&bounded=1&viewbox=-46.85,-23.35,-46.35,-23.85"
            
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read().decode())
                for result in data:
                    lat, lon = float(result["lat"]), float(result["lon"])
                    if LAT_MIN < lat < LAT_MAX and LNG_MIN < lon < LNG_MAX:
                        return lat, lon, "Real"
            time.sleep(1.2)
        except:
            pass
        
    return None, None, None

def formatar_entrega(valor):
    """Formata a data de entrega para exibição."""
    if not valor:
        return ""
    val = str(valor).strip()
    # Se for datetime, extrair apenas ano
    if "00:00:00" in val:
        try:
            from datetime import datetime
            dt = datetime.strptime(val, "%Y-%m-%d %H:%M:%S")
            meses = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
                     "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
            return f"{meses[dt.month - 1]}/{dt.year}"
        except:
            return val
    return val

def verificar_imagem(nome_arquivo):
    """Verifica se a imagem existe na pasta e retorna o caminho."""
    if not nome_arquivo:
        return ""
    caminho = os.path.join(PASTA_IMAGENS, nome_arquivo.strip())
    if os.path.exists(caminho):
        return f"{PASTA_IMAGENS}/{nome_arquivo.strip()}"
    else:
        print(f"  ⚠️  Imagem não encontrada: {caminho}")
        return ""

def main():
    print("=" * 55)
    print("  MT Parceiros — CMS Sync v1.0")
    print("  Excel → Site (Mapa + Listagens + Imagens)")
    print("=" * 55)

    if not os.path.exists(ARQUIVO_XLSX):
        print(f"\n❌ Arquivo '{ARQUIVO_XLSX}' não encontrado.")
        sys.exit(1)

    cache = carregar_cache()
    wb = openpyxl.load_workbook(ARQUIVO_XLSX)
    ws = wb.active
    empreendimentos = []
    erros = 0
    linhas = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
    total = sum(1 for r in linhas if r[COL_NOME])
    print(f"📋 Processando {total} empreendimento(s)...\n")

    for i, row in enumerate(linhas):
        nome = str(row[COL_NOME] or "").strip()
        if not nome:
            continue
        
        endereco = str(row[COL_END] or "").strip().replace("\n", " ")
        bairro   = str(row[COL_BAIRRO] or "").strip()
        
        # Tentar extrair CEP do endereço ou bairro se houver (padrão 00000-000)
        match_cep = re.search(r'(\d{5}-?\d{3})', f"{endereco} {bairro}")
        cep_encontrado = match_cep.group(1) if match_cep else None
        cep_info = consultar_viacep(cep_encontrado) if cep_encontrado else None

        print(f"🔍 [{i+1}/{total}] {nome}")
        if cep_info:
            print(f"  📮 (ViaCEP) {cep_info['logradouro']} - {cep_info['bairro']}")

        # Geocodificação
        lat, lng = None, None
        if endereco in cache:
            lat, lng = cache[endereco]["lat"], cache[endereco]["lng"]
            print(f"  🧊 (Cache) {lat:.5f}, {lng:.5f}")
        else:
            # Tentar Google primeiro se houver chave (legado do plano anterior, mantido se o user conseguir depois)
            if GOOGLE_KEY:
                lat, lng, tipo = geocodificar_google(endereco, bairro)
            
            # Backup ou Fallback para Nominatim (com info extra do CEP)
            if not lat:
                lat, lng, tipo = geocodificar_nominatim(endereco, bairro, cep_info)
                
            if lat:
                print(f"  📍 ({tipo}) {lat:.5f}, {lng:.5f}")
                cache[endereco] = {"lat": lat, "lng": lng}
                salvar_cache(cache)
            else:
                print(f"  ❌ Falha no endereço: {endereco}")
                erros += 1
                continue

        # Verificar imagem
        imagem_nome = str(row[COL_IMAGEM] or "").strip()
        imagem_path = verificar_imagem(imagem_nome)
        if imagem_path:
            print(f"  🖼️  {imagem_path}")

        # Montar objeto completo
        empreendimentos.append({
            "nome":          nome.replace("\n", " "),
            "bairro":        bairro,
            "endereco":      endereco,
            "area":          str(row[COL_AREA] or ""),
            "quartos":       str(row[COL_QUARTOS] or ""),
            "diferenciais":  str(row[COL_DIFER] or ""),
            "lazer":         str(row[COL_LAZER] or ""),
            "preco":         str(row[COL_PRECO] or ""),
            "entrega":       formatar_entrega(row[COL_ENTREGA]),
            "imagem":        imagem_path,
            "lat":           lat,
            "lng":           lng
        })

    # Gerar arquivo JS
    with open(ARQUIVO_JS, "w", encoding="utf-8") as f:
        f.write("var EMPREENDIMENTOS = ")
        json.dump(empreendimentos, f, ensure_ascii=False, indent=2)
        f.write(";")

    # Relatório final
    print("\n" + "=" * 55)
    print(f"✅ {len(empreendimentos)} empreendimentos sincronizados!")
    if erros > 0:
        print(f"⚠️  {erros} endereço(s) com falha de geocodificação.")
    print(f"📄 Arquivo gerado: {ARQUIVO_JS}")
    print(f"🗺️  Mapa atualizado com {len(empreendimentos)} pins.")
    print("=" * 55)

if __name__ == "__main__":
    main()
