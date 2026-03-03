import openpyxl
import json

def analyze_excel(path):
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        report = {
            "sheets": [],
            "samples": {}
        }
        
        for name in wb.sheetnames:
            sheet = wb[name]
            report["sheets"].append({
                "name": name,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column
            })
            
            # Get first 30 rows and 15 columns
            rows = []
            for row in sheet.iter_rows(max_row=30, max_col=15, values_only=True):
                processed_row = []
                for cell in row:
                    if hasattr(cell, 'isoformat'):
                        processed_row.append(cell.isoformat())
                    else:
                        processed_row.append(str(cell) if cell is not None else None)
                rows.append(processed_row)
            report["samples"][name] = rows
            
        print(json.dumps(report, indent=2))
    except Exception as e:
        print(f"Error: {str(e)}")

analyze_excel(r"C:\Users\Marcos.PC_M1\Downloads\Laudo_MT_Parceiros_ruiva.xlsx")
