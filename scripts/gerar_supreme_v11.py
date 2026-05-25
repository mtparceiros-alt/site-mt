import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def gerar_supreme_v11(path_input, path_output):
    try:
        # 1. CARREGAR O MOTOR (v5_charts)
        wb = openpyxl.load_workbook(path_input, data_only=False)
        
        # 2. CONFIGURAR ABAS
        if 'Dashboard MT' in wb.sheetnames: del wb['Dashboard MT']
        ws = wb.create_sheet("Dashboard MT", 0) # Criar como primeira aba
        ws.sheet_view.showGridLines = False

        # Renomear/Configurar aba de fluxo (se existir v5)
        if 'MT Parceiros' in wb.sheetnames:
            ws_flow = wb['MT Parceiros']
            ws_flow.title = "Fluxo Detalhado"
        
        # --- PALETA SUPREME BI ---
        COLOR_BG = "0a0f1c"       
        COLOR_CARD = "121a2b"     
        COLOR_GOLD = "d4af37"     
        COLOR_ORANGE = "f35525"   
        COLOR_WHITE = "ffffff"
        COLOR_GRAY = "b0b8c9"
        
        COLOR_PURPLE = "8a2be2"
        COLOR_BLUE = "1e90ff"
        COLOR_GREEN = "32cd32"

        # 3. PINTAR FUNDO DASHBOARD
        fill_bg = PatternFill(start_color=COLOR_BG, end_color=COLOR_BG, fill_type="solid")
        for row in range(1, 101):
            for col in range(1, 40):
                ws.cell(row=row, column=col).fill = fill_bg

        # 4. GRID DE ESTABILIDADE (2 QUADRANTES LATERAIS A-E e G-J)
        cols_width = {'A':3, 'B':22, 'C':18, 'D':14, 'E':14, 'F':6, 'G':22, 'H':14, 'I':14, 'J':18, 'K':3}
        for col, width in cols_width.items(): ws.column_dimensions[col].width = width

        # 5. HEADER (B2:J6)
        ws.merge_cells('B2:J6')
        header = ws['B2']
        header.value = "MT PARCEIROS • DOSSIÊ SUPREMO BI 2026"
        header.font = Font(name='Arial', size=28, bold=True, color=COLOR_GOLD)
        header.alignment = Alignment(horizontal='center', vertical='center')

        # --- FUNÇÃO HELPER PARA CARDS ---
        fill_card = PatternFill(start_color=COLOR_CARD, end_color=COLOR_CARD, fill_type="solid")
        border_gold = Border(left=Side(style='thin', color=COLOR_GOLD), right=Side(style='thin', color=COLOR_GOLD),
                            top=Side(style='thin', color=COLOR_GOLD), bottom=Side(style='thin', color=COLOR_GOLD))

        def desenhar_quadrante(ws, r_start, c_start, title):
            # Título do Módulo
            ws.cell(row=r_start-1, column=c_start).value = title
            ws.cell(row=r_start-1, column=c_start).font = Font(name='Arial', size=14, bold=True, color=COLOR_GOLD)
            
            for r in range(r_start, r_start+15):
                for c in range(c_start, c_start+4):
                    ws.cell(row=r, column=c).fill = fill_card
                    # Borda externa do Quadrante
                    if r == r_start: ws.cell(row=r, column=c).border = Border(top=Side(style='thin', color=COLOR_GOLD))
                    if r == r_start+14: ws.cell(row=r, column=c).border = Border(bottom=Side(style='thin', color=COLOR_GOLD))
                    if c == c_start: ws.cell(row=r, column=c).border = Border(left=Side(style='thin', color=COLOR_GOLD))
                    if c == c_start+3: ws.cell(row=r, column=c).border = Border(right=Side(style='thin', color=COLOR_GOLD))

        # 6. QUADRANTE 1: PERFIL & RECURSOS (B10:E24)
        desenhar_quadrante(ws, 10, 2, "👤 1. PERFIL & RECURSOS")
        inputs = [("B12", "Nome Completo:", "C12"), ("B14", "Renda Familiar:", "C14"),
                  ("B16", "Recursos (Dinheiro):", "C16"), ("B18", "Saldo FGTS:", "C18"),
                  ("B20", "Total de Dívidas:", "C20")]
        for lb, t, v in inputs:
            ws[lb].value = t
            ws[lb].font = Font(name='Arial', size=11, color=COLOR_WHITE)
            ws[v].font = Font(name='Arial', size=12, bold=True, color=COLOR_WHITE)
            ws[v].border = Border(bottom=Side(style='thin', color=COLOR_GOLD))

        # 7. QUADRANTE 2: ANÁLISE CAIXA (G10:J24)
        desenhar_quadrante(ws, 10, 7, "🏦 2. ANÁLISE CAIXA")
        ws['G12'].value = "ENQUADRAMENTO:"
        ws['I12'].value = "=IF(C14<=2100, \"FAIXA 1\", IF(C14<=4400, \"FAIXA 2\", \"FAIXA 3\"))"
        ws['I12'].font = Font(name='Arial', size=12, bold=True, color=COLOR_WHITE)

        ws['G15'].value = "CRÉDITO LIBERADO:"
        ws['I15'].font = Font(name='Arial', size=14, bold=True, color=COLOR_WHITE)
        
        ws['G18'].value = "SUBSÍDIO ESTIMADO:"
        ws['I18'].font = Font(name='Arial', size=14, bold=True, color=COLOR_ORANGE)

        ws['G21'].value = "PARCELA PROJETADA:"
        ws['I21'].font = Font(name='Arial', size=14, bold=True, color=COLOR_GREEN)

        # 8. QUADRANTE 3: LAUDO DE PODER (B30:E44)
        desenhar_quadrante(ws, 30, 2, "🏆 3. LAUDO DE CRÉDITO")
        ws['B34'].value = "PODER DE COMPRA TOTAL:"
        ws['B34'].font = Font(name='Arial', size=12, bold=True, color=COLOR_GRAY)
        ws.merge_cells('B35:E40')
        ws['B35'].value = "=ROUND(I15+I18+C16+C18, 0)" # Crédito + Subsídio + Entrada + FGTS
        ws['B35'].font = Font(name='Arial', size=42, bold=True, color=COLOR_ORANGE)
        ws['B35'].alignment = Alignment(horizontal='center', vertical='center')

        # 9. QUADRANTE 4: PLANO DE JORNADA (G30:J44) - Grade 2x2 Interna
        desenhar_quadrante(ws, 30, 7, "📍 4. PLANO DE JORNADA")
        
        def mini_card_2x2(ws, r, c, label, border_color):
            border_side = Side(style='thick', color=border_color)
            for ri in range(r, r+6):
                for ci in range(c, c+2):
                    ws.cell(row=ri, column=ci).fill = fill_card
                    if ci == c: ws.cell(row=ri, column=ci).border = Border(left=border_side)
            ws.cell(row=r, column=c).value = label
            ws.cell(row=r, column=c).font = Font(name='Arial', size=8, bold=True, color=COLOR_GRAY)
            ws.merge_cells(start_row=r+1, start_column=c, end_row=r+3, end_column=c+1)
            val_cell = ws.cell(row=r+1, column=c)
            val_cell.font = Font(name='Arial', size=14, bold=True, color=COLOR_WHITE)
            val_cell.alignment = Alignment(horizontal='center', vertical='center')
            return val_cell

        # 2x2 MiniCards
        mini_card_2x2(ws, 32, 7, "ENQUADRAMENTO", COLOR_PURPLE).value = "HIS-2 / MCMV" # G33
        mini_card_2x2(ws, 32, 9, "ENTRADA (36x)", COLOR_ORANGE).value = "R$ 0,00" # I33
        mini_card_2x2(ws, 39, 7, "EVOLUÇÃO OBRA", COLOR_BLUE).value = "R$ 0,00" # G40
        mini_card_2x2(ws, 39, 9, "POS-CHAVES", COLOR_GREEN).value = "R$ 0,00" # I40

        # 10. RODAPÉ SUPREMO (A50:AZ100)
        ws.merge_cells('B50:J53')
        ws['B50'].value = "📲 CONTINUAR ATENDIMENTO EXCLUSIVO VIA WHATSAPP"
        ws['B50'].font = Font(name='Arial', size=18, bold=True, color=COLOR_WHITE)
        ws['B50'].fill = PatternFill(start_color=COLOR_ORANGE, end_color=COLOR_ORANGE, fill_type="solid")
        ws['B50'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B50'].hyperlink = "https://wa.me/5511960364355"

        # Ocultar outras abas (exceto Fluxo Detalhado)
        for s in wb.worksheets:
            if s.title not in ["Dashboard MT", "Fluxo Detalhado"]:
                s.sheet_state = "hidden"

        wb.save(path_output)
        print(f"Sucesso v11 (2x2 BI): {path_output}")

    except Exception as e: print(f"Erro v11: {e}")

if __name__ == "__main__":
    gerar_supreme_v11(
        "c:/Users/Marcos.PC_M1/Documents/site_mt/assets/docs/template_mt_parceiros_v5_charts.xlsx",
        "c:/Users/Marcos.PC_M1/Documents/site_mt/assets/docs/template_mt_parceiros_v11_supreme.xlsx"
    )
