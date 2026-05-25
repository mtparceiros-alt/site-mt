import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def restaurar_cerebro_v10(path_input, path_output):
    try:
        # 1. CARREGAR O MOTOR ORIGINAL (v5_charts tem toda a lógica)
        wb = openpyxl.load_workbook(path_input, data_only=False)
        
        # 2. SELECIONAR A ABA PRINCIPAL (MT Parceiros) E LIMPAR MERGES
        ws = wb['MT Parceiros']
        ws.sheet_view.showGridLines = False
        
        merges = list(ws.merged_cells.ranges)
        for m in merges: ws.unmerge_cells(str(m))

        # --- PALETA GROK SUPREME ---
        COLOR_BG = "0a0f1c"       
        COLOR_CARD = "121a2b"     
        COLOR_ORANGE = "f35525"   
        COLOR_GOLD = "d4af37"     
        COLOR_PURPLE = "8a2be2"
        COLOR_BLUE = "1e90ff"
        COLOR_GREEN = "32cd32"
        COLOR_WHITE = "ffffff"
        COLOR_GRAY = "b0b8c9"

        # 3. PINTAR FUNDO (Preservando as fórmulas nas células distantes se houver)
        fill_bg = PatternFill(start_color=COLOR_BG, end_color=COLOR_BG, fill_type="solid")
        for row in range(1, 201):
            for col in range(1, 40):
                ws.cell(row=row, column=col).fill = fill_bg

        # 4. CONFIGURAÇÃO DE COLUNAS (Otimizado v10)
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['J'].width = 20

        # 5. HEADER SUPREME (B2:J5)
        ws['B2'].value = "MT PARCEIROS • RELATÓRIO DE VIABILIDADE SUPREMO"
        ws['B2'].font = Font(name='Arial', size=26, bold=True, color=COLOR_GOLD)
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('B2:J5')

        # 6. RESTAURAR CARDS ORIGINAIS (PERFIL E LAUDO)
        fill_card = PatternFill(start_color=COLOR_CARD, end_color=COLOR_CARD, fill_type="solid")
        for r in range(10, 26):
            for c in range(2, 6): ws.cell(row=r, column=c).fill = fill_card
            for c in range(7, 11): ws.cell(row=r, column=c).fill = fill_card
        
        ws['B10'].value = "👤 PERFIL DO CLIENTE"
        ws['G10'].value = "📊 LAUDO TÉCNICO MCMV"
        for p in ['B10', 'G10']: ws[p].font = Font(name='Arial', size=14, bold=True, color=COLOR_GOLD)

        # Labels e Vínculos de Fórmulas (Injetando nos nomes corretos)
        # C14 (Renda), D27 (Poder de Compra original da v5)
        ws['G17'].value = "PODER DE COMPRA:"
        ws['G18'].value = "=D27" # APONTANDO PARA A FÓRMULA ORIGINAL
        ws['G18'].font = Font(name='Arial', size=36, bold=True, color=COLOR_ORANGE)
        ws['G18'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('G18:J22')

        # 7. MÓDULO 3: PLANO DE FLUXO (Cards)
        def criar_card_v10(ws, row, title, color):
            side = Side(style='thick', color=color)
            for r in range(row, row+5):
                for c in range(2, 11):
                    ws.cell(row=r, column=c).fill = fill_card
                    if c == 2: ws.cell(row=r, column=c).border = Border(left=side)
            ws.cell(row=row, column=3).value = title
            ws.cell(row=row, column=3).font = Font(name='Arial', size=10, bold=True, color=COLOR_GRAY)
            ws.merge_cells(start_row=row+1, start_column=3, end_row=row+2, end_column=9)
            ws.cell(row=row+1, column=3).font = Font(name='Arial', size=24, bold=True, color=COLOR_WHITE)

        criar_card_v10(ws, 30, "ENQUADRAMENTO (SP 2026)", COLOR_PURPLE) # C31
        criar_card_v10(ws, 37, "ENTRADA PARCELADA (CONSTRUTORA)", COLOR_ORANGE) # C38
        criar_card_v10(ws, 44, "EVOLUÇÃO DE OBRA (MÉDIA)", COLOR_BLUE) # C45
        criar_card_v10(ws, 51, "PRESTAÇÃO MENSAL (PÓS-CHAVES)", COLOR_GREEN) # C52

        # 8. RODAPÉ E BOTÃO
        ws.merge_cells('B160:J163')
        ws['B160'].value = "📲 CONTINUAR ATENDIMENTO EXCLUSIVO VIA WHATSAPP"
        ws['B160'].font = Font(name='Arial', size=18, bold=True, color=COLOR_WHITE)
        ws['B160'].fill = PatternFill(start_color=COLOR_ORANGE, end_color=COLOR_ORANGE, fill_type="solid")
        ws['B160'].alignment = Alignment(horizontal='center', vertical='center')

        ws.freeze_panes = "A9"
        wb.save(path_output)
        print(f"Sucesso v10 (Fórmulas Ativas): {path_output}")

    except Exception as e: print(f"Erro Crítico v10: {e}")

if __name__ == "__main__":
    restaurar_cerebro_v10(
        "c:/Users/Marcos.PC_M1/Documents/site_mt/assets/docs/template_mt_parceiros_v5_charts.xlsx",
        "c:/Users/Marcos.PC_M1/Documents/site_mt/assets/docs/template_mt_parceiros_v10_supreme.xlsx"
    )
