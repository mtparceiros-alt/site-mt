import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def gerar_dossie_v12(path_input, path_output):
    """
    ✒️ CARTA DE CRÉDITO SUPREMO (v12.7) - Simplificada
    Este script lê a planilha bruta do simulador financeiro e a transforma 
    em um Dossiê de Viabilidade de Investimento de Alto Padrão.
    """
    try:
        # [PASSO 1] CARREGAMENTO DA MÁQUINA DE CÁLCULO
        wb = openpyxl.load_workbook(path_input, data_only=False)
        
        # Limpeza de abas dashboard antigas
        for sn in ['Dashboard MT', 'Dossiê de Crédito', 'Painel Visual', 'Plano de Ação']:
            if sn in wb.sheetnames: del wb[sn]
            
        # Criação da tela de pintura (Aba Principal)
        ws = wb.create_sheet("Dossiê de Crédito", 0)
        ws.sheet_view.showGridLines = False

        # Renomeio da aba velha (Hospeda os dados para as fórmulas)
        if 'MT Parceiros' in wb.sheetnames:
            tech_s = wb['MT Parceiros']
            tech_s.title = "Fluxo Detalhado"
            tech_s.sheet_state = "hidden" 

        # =========================================================
        # [PASSO 2] IDENTIDADE VISUAL E CORES
        # =========================================================
        COLOR_BG = "0a0f1c"       # Graphite Fundo
        COLOR_CARD = "131A2A"     # Fundo dos Cards Principais
        COLOR_CARD_LIGHT = "192338" # Fundo Cards Secundários
        COLOR_WHITE = "FFFFFF"
        COLOR_TEXT = "A1A9B9"     # Texto secundário
        COLOR_ACCENT = "f35525"   # Laranja MT
        COLOR_GREEN = "10B981"    # Verde
        COLOR_GOLD = "d4af37"     # Ouro 
        
        # Pinta o fundo da folha
        fill_bg = PatternFill(start_color=COLOR_BG, end_color=COLOR_BG, fill_type="solid")
        for row in range(1, 100):
            for col in range(1, 20):
                ws.cell(row=row, column=col).fill = fill_bg

        def block(r1, c1, r2, c2, color):
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for r in range(r1, r2+1):
                for col in range(c1, c2+1):
                    ws.cell(row=r, column=col).fill = fill
                    
        # Margens laterais
        cols_width = {'A':4, 'B':4, 'C':15, 'D':15, 'E':15, 'F':15, 'G':15, 'H':15, 'I':15, 'J':15, 'K':4}
        for char, width in cols_width.items():
            ws.column_dimensions[char].width = width

        # =========================================================
        # CABEÇALHO 
        # =========================================================
        # Título à esquerda
        ws.cell(row=2, column=3, value="DOSSIÊ DE VIABILIDADE E CRÉDITO").font = Font(name='Arial', size=18, bold=True, color=COLOR_WHITE)
        ws.merge_cells('C2:G3')
        ws.cell(row=2, column=3).alignment = Alignment(horizontal='left', vertical='center')

        # Status à direita (H2 em diante)
        ws.cell(row=2, column=8, value="✔ APROVADO (76 PTS)").font = Font(name='Arial', size=11, bold=True, color=COLOR_GREEN)
        ws.merge_cells('H2:J3')
        ws.cell(row=2, column=8).alignment = Alignment(horizontal='right', vertical='center')

        # Linha do Cliente (Nome)
        ws.cell(row=4, column=3, value="MARCOS HENRIQUE ANDRADE DE MEDEIROS | 30 ANOS").font = Font(name='Arial', size=8, bold=False, color=COLOR_GOLD)
        ws.merge_cells('C4:J4')
        ws.cell(row=4, column=3).alignment = Alignment(horizontal='left', vertical='center')

        # [FAIXA 1] PODER DE COMPRA
        block(6, 3, 11, 10, COLOR_CARD)
        
        ws.cell(row=8, column=4, value="CAPACIDADE TOTAL DE COMPRA LIBERADA PARA NEGOCIAÇÃO").font = Font(name='Arial', size=10, bold=True, color=COLOR_TEXT)
        ws.merge_cells('D8:I8')
        ws.cell(row=8, column=4).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=9, column=4, value="").font = Font(name='Arial', size=38, bold=True, color=COLOR_ACCENT)
        ws.merge_cells('D9:I10')
        ws.cell(row=9, column=4).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=9, column=4).number_format = '"R$ "#,##0.00'

        # [FAIXA 2] POTENCIAL
        block(13, 3, 18, 6, COLOR_CARD_LIGHT)
        ws.cell(row=14, column=3, value="DADOS DE CRÉDITO IA").font = Font(name='Arial', size=11, bold=True, color=COLOR_WHITE)
        ws.merge_cells('C14:F14')
        ws.cell(row=14, column=3).alignment = Alignment(horizontal='center')
        
        ws.cell(row=16, column=3, value=" CRÉDITO CAIXA:").font = Font(name='Arial', size=10, color=COLOR_TEXT)
        ws.cell(row=16, column=5, value="='Fluxo Detalhado'!D27").font = Font(name='Arial', size=10, bold=True, color=COLOR_WHITE)
        ws.cell(row=16, column=5).number_format = '"R$ "#,##0.00'

        ws.cell(row=17, column=3, value=" SUBSÍDIO (Governo):").font = Font(name='Arial', size=10, color=COLOR_TEXT)
        ws.cell(row=17, column=5, value="='Fluxo Detalhado'!D29").font = Font(name='Arial', size=10, bold=True, color=COLOR_GREEN)
        ws.cell(row=17, column=5).number_format = '"R$ "#,##0.00'

        ws.cell(row=18, column=3, value=" SALDO FGTS:").font = Font(name='Arial', size=10, color=COLOR_TEXT)
        ws.cell(row=18, column=5, value="='Fluxo Detalhado'!C14").font = Font(name='Arial', size=10, bold=True, color=COLOR_WHITE)
        ws.cell(row=18, column=5).number_format = '"R$ "#,##0.00'

        block(13, 7, 18, 10, COLOR_CARD)
        ws.cell(row=14, column=8, value="O SEU INVESTIMENTO HOJE").font = Font(name='Arial', size=11, bold=True, color=COLOR_WHITE)
        ws.merge_cells('H14:I14')
        ws.cell(row=14, column=8).alignment = Alignment(horizontal='center')

        ws.cell(row=16, column=8, value="Entrada Disponível / Acordada").font = Font(name='Arial', size=9, color=COLOR_TEXT)
        ws.merge_cells('H16:I16')
        ws.cell(row=16, column=8).alignment = Alignment(horizontal='center')
        
        ws.cell(row=17, column=8, value="='Fluxo Detalhado'!C13").font = Font(name='Arial', size=18, bold=True, color=COLOR_WHITE)
        ws.merge_cells('H17:I18')
        ws.cell(row=17, column=8).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=17, column=8).number_format = '"R$ "#,##0.00'

        # [FAIXA 3] FLUXO 36 MESES
        block(20, 3, 20, 10, COLOR_BG) 
        block(21, 3, 28, 10, COLOR_CARD_LIGHT)

        ws.cell(row=22, column=3, value="💳 PLANO DE FLUXO DE PAGAMENTO (CONSTRUTORA + BANCO)").font = Font(name='Arial', size=11, bold=True, color=COLOR_ACCENT)
        ws.merge_cells('C22:J22')
        ws.cell(row=22, column=3).alignment = Alignment(horizontal='center', vertical='center')

        # Sub-Cards
        block(24, 3, 27, 4, "0d1929")
        ws.cell(row=24, column=3, value="ENQUADRAMENTO / ITBI").font = Font(name='Arial', size=7, bold=True, color=COLOR_TEXT)
        ws.merge_cells('C24:D24')
        ws.cell(row=24, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=25, column=3, value="='Fluxo Detalhado'!D33").font = Font(name='Arial', size=10, bold=True, color=COLOR_WHITE)
        ws.merge_cells('C25:D26')
        ws.cell(row=25, column=3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=27, column=3, value='="ITBI " & \'Fluxo Detalhado\'!D36').font = Font(name='Arial', size=8, bold=True, color=COLOR_GREEN)
        ws.merge_cells('C27:D27')
        ws.cell(row=27, column=3).alignment = Alignment(horizontal='center')

        block(24, 5, 27, 6, "0d1929")
        ws.cell(row=24, column=5, value="ENTRADA PARCELADA (CONSTRUTORA)").font = Font(name='Arial', size=7, bold=True, color=COLOR_TEXT)
        ws.merge_cells('E24:F24')
        ws.cell(row=24, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=25, column=5, value="='Fluxo Detalhado'!D31").font = Font(name='Arial', size=16, bold=True, color=COLOR_ACCENT)
        ws.merge_cells('E25:F26')
        ws.cell(row=25, column=5).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=25, column=5).number_format = '"R$ "#,##0.00'
        ws.cell(row=27, column=5, value="Período: 36 meses").font = Font(name='Arial', size=8, color=COLOR_TEXT)
        ws.merge_cells('E27:F27')
        ws.cell(row=27, column=5).alignment = Alignment(horizontal='center')

        block(24, 7, 27, 8, "0d1929")
        ws.cell(row=24, column=7, value="EVOLUÇÃO DE OBRA (MÉDIA)").font = Font(name='Arial', size=7, bold=True, color=COLOR_TEXT)
        ws.merge_cells('G24:H24')
        ws.cell(row=24, column=7).alignment = Alignment(horizontal='center')
        ws.cell(row=25, column=7, value="").font = Font(name='Arial', size=16, bold=True, color="00bcd4")
        ws.merge_cells('G25:H26')
        ws.cell(row=25, column=7).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=25, column=7).number_format = '"R$ "#,##0.00'
        ws.cell(row=27, column=7, value="Conforme a obra avança").font = Font(name='Arial', size=8, color=COLOR_TEXT)
        ws.merge_cells('G27:H27')
        ws.cell(row=27, column=7).alignment = Alignment(horizontal='center')

        block(24, 9, 27, 10, "0d1929")
        ws.cell(row=24, column=9, value="PRESTAÇÃO MENSAL (PÓS-CHAVES)").font = Font(name='Arial', size=7, bold=True, color=COLOR_TEXT)
        ws.merge_cells('I24:J24')
        ws.cell(row=24, column=9).alignment = Alignment(horizontal='center')
        ws.cell(row=25, column=9, value="='Fluxo Detalhado'!D30").font = Font(name='Arial', size=16, bold=True, color=COLOR_GREEN)
        ws.merge_cells('I25:J26')
        ws.cell(row=25, column=9).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=25, column=9).number_format = '"R$ "#,##0.00'
        ws.cell(row=27, column=9, value="Financiamento definitivo").font = Font(name='Arial', size=8, color=COLOR_TEXT)
        ws.merge_cells('I27:J27')
        ws.cell(row=27, column=9).alignment = Alignment(horizontal='center')

        # [FAIXA 4] VALORIZAÇÃO
        block(29, 3, 29, 10, COLOR_BG) 
        block(30, 3, 38, 10, COLOR_CARD_LIGHT)

        ws.cell(row=31, column=4, value="📈 AUMENTO DE PATRIMÔNIO (PROJEÇÃO NA CHAVE)").font = Font(name='Arial', size=12, bold=True, color=COLOR_GREEN)
        ws.merge_cells('D31:I31')
        ws.cell(row=31, column=4).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=32, column=3, value="").font = Font(name='Arial', size=9, italic=True, color=COLOR_GOLD)
        ws.merge_cells('C32:J32')
        ws.cell(row=32, column=3).alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=33, column=3, value="VALOR DO IMÓVEL (HOJE)").font = Font(name='Arial', size=10, color=COLOR_TEXT)
        ws.merge_cells('C33:F33')
        ws.cell(row=33, column=3).alignment = Alignment(horizontal='center')

        ws.cell(row=34, column=3, value="").font = Font(name='Arial', size=18, bold=True, color=COLOR_WHITE)
        ws.merge_cells('C34:F36')
        ws.cell(row=34, column=3).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=34, column=3).number_format = '"R$ "#,##0.00'

        ws.cell(row=33, column=7, value="VALOR NA ENTREGA (MÉDIA +30%)").font = Font(name='Arial', size=10, color=COLOR_TEXT)
        ws.merge_cells('G33:J33')
        ws.cell(row=33, column=7).alignment = Alignment(horizontal='center')

        ws.cell(row=34, column=7, value="='Dossiê de Crédito'!C34 * 1.30").font = Font(name='Arial', size=24, bold=True, color=COLOR_GREEN)
        ws.merge_cells('G34:J36')
        ws.cell(row=34, column=7).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=34, column=7).number_format = '"R$ "#,##0.00'

        # [FAIXA 5] BLINDAGEM
        ws.cell(row=40, column=3, value="🛡️ A BLINDAGEM 360° MT PARCEIROS").font = Font(name='Arial', size=12, bold=True, color=COLOR_GOLD)
        ws.merge_cells('C40:J40')
        ws.cell(row=40, column=3).alignment = Alignment(horizontal='center', vertical='center')

        selos = [
            "🏦 1. APROVAÇÃO BANCÁRIA\nTodo limite é analisado via API direto antes da emissão do contrato.",
            "⚖️ 2. ANÁLISE JURÍDICA\nNossos advogados validam a idoneidade e cláusulas perigosas.",
            "🏗️ 3. ACOMPANHAMENTO\nNão abandonamos você. Nossa equipe acompanha as medições de obra.",
            "👷 4. LAUDO DE ENGENHARIA\nSelo MT: O fechamento garante que nossa engenheira fará sua Vistoria Final."
        ]

        col_idx = 3
        for selo in selos:
            ws.cell(row=42, column=col_idx, value=selo).font = Font(name='Arial', size=9, color=COLOR_WHITE)
            ws.merge_cells(start_row=42, start_column=col_idx, end_row=46, end_column=col_idx+1)
            ws.cell(row=42, column=col_idx).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            block(42, col_idx, 46, col_idx+1, COLOR_CARD)
            col_idx += 2 

        ws.cell(row=49, column=3, value="ESTE DOSSIÊ É ESTRITAMENTE CONFIDENCIAL E TEM FINS DE PLANEJAMENTO.").font = Font(name='Arial', size=7, color="475569")
        ws.merge_cells('C49:J49')
        ws.cell(row=49, column=3).alignment = Alignment(horizontal='center')

        # Ocultar abas
        for s in wb.worksheets:
            if s.title != "Dossiê de Crédito":
                s.sheet_state = "hidden"

        wb.save(path_output)
        print(f"✅ SUCESSO: Dossiê Supremo v12.7 (Simplificado) Gerado em -> {path_output}")

    except Exception as e: 
        import traceback
        print(f"❌ ERRO: {str(e)}")
        traceback.print_exc()

if __name__ == "__main__":
    gerar_dossie_v12(
        "c:/Users/Marcos.PC_M1/Documents/site_mt/assets/docs/template_mt_parceiros_v5_charts.xlsx",
        "c:/Users/Marcos.PC_M1/Documents/site_mt/assets/docs/template_mt_parceiros_v12_dossie.xlsx"
    )
