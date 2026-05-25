import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def gerar_elite_v7_completo(path_output):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MT Elite Dash"
        ws.sheet_view.showGridLines = False

        # --- PALETA GROK PREMIUM ---
        COLOR_BG = "0a0f1c"       
        COLOR_CARD = "121a2b"     
        COLOR_ORANGE = "f35525"   
        COLOR_GOLD = "d4af37"     
        COLOR_WHITE = "ffffff"
        COLOR_GRAY = "b0b8c9"

        # 1. Pintar fundo total (A1:AZ200)
        fill_bg = PatternFill(start_color=COLOR_BG, end_color=COLOR_BG, fill_type="solid")
        for row in range(1, 401):
            for col in range(1, 40):
                ws.cell(row=row, column=col).fill = fill_bg

        # 2. CONFIGURAÇÃO DE COLUNAS (GUTTERS)
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 5 
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 20
        ws.column_dimensions['K'].width = 3

        # 3. HEADER PREMIUM (B2:J8)
        ws.merge_cells('B2:J5')
        header = ws['B2']
        header.value = "MT PARCEIROS • RELATÓRIO INTELIGENTE DE VIABILIDADE"
        header.font = Font(name='Arial', size=26, bold=True, color=COLOR_GOLD)
        header.alignment = Alignment(horizontal='center', vertical='center')
        
        ws['B6'].value = "📍 mtparceiros@gmail.com  |  📞 (11) 96036-4355"
        ws['B6'].font = Font(name='Arial', size=11, color=COLOR_GRAY)
        
        ws['I6'].value = "DATA:"
        ws['I6'].font = Font(name='Arial', size=11, color=COLOR_GRAY)
        ws['J6'].value = "=TEXT(TODAY(),\"dd/mm/yyyy\")"
        ws['J6'].font = Font(name='Arial', size=11, bold=True, color=COLOR_WHITE)

        # 4. CARD 1: PERFIL FINANCEIRO (B9:E30)
        fill_card = PatternFill(start_color=COLOR_CARD, end_color=COLOR_CARD, fill_type="solid")
        for r in range(9, 31):
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill_card
        
        ws['B9'].value = "👤 IDENTIFICAÇÃO DO CLIENTE"
        ws['B9'].font = Font(name='Arial', size=14, bold=True, color=COLOR_GOLD)
        
        ws['B11'].value = "Nome do Cliente:"
        ws['C11'].value = "Cliente MT" # JS Injetará aqui
        
        ws['B14'].value = "📥 DADOS DE ENTRADA"
        ws['B14'].font = Font(name='Arial', size=14, bold=True, color=COLOR_GOLD)
        
        labels = [("B16", "Renda Bruta (R$):", "C16"), 
                  ("B17", "Entrada Dinheiro (R$):", "C17"),
                  ("B18", "Saldo FGTS (R$):", "C18"),
                  ("B19", "Carteira Assinada?", "C19"),
                  ("B20", "Dívidas Atuais (R$):", "C20")]
        
        for lb_pos, text, val_pos in labels:
            ws[lb_pos].value = text
            ws[lb_pos].font = Font(name='Arial', size=11, color=COLOR_WHITE)
            ws[val_pos].font = Font(name='Arial', size=12, bold=True, color=COLOR_WHITE)
            ws[val_pos].border = Border(bottom=Side(style='thin', color=COLOR_GOLD))

        # 5. CARD 2: RESULTADOS (G14:J30)
        for r in range(14, 31):
            for c in range(7, 11):
                ws.cell(row=r, column=c).fill = fill_card

        ws['G14'].value = "📊 PARECER DE VIABILIDADE"
        ws['G14'].font = Font(name='Arial', size=14, bold=True, color=COLOR_GOLD)
        
        ws['G16'].value = "PODER DE COMPRA:"
        ws['G16'].font = Font(name='Arial', size=11, color=COLOR_WHITE)
        
        ws.merge_cells('G17:J21')
        ws['G17'].value = "0" # JS Injetará aqui
        ws['G17'].font = Font(name='Arial', size=36, bold=True, color=COLOR_ORANGE)
        ws['G17'].alignment = Alignment(horizontal='center', vertical='center')

        ws['G23'].value = "STATUS:"
        ws['G23'].font = Font(name='Arial', size=11, color=COLOR_WHITE)
        ws['I23'].value = "EM ANÁLISE" # JS Injetará aqui
        ws['I23'].font = Font(name='Arial', size=14, bold=True, color=COLOR_GOLD)

        # 6. SCORE VISUAL (B33:J38)
        for r in range(33, 39):
            for c in range(2, 11):
                ws.cell(row=r, column=c).fill = fill_card
        
        ws['B33'].value = "🏥 SCORE DE JORNADA MT"
        ws['B33'].font = Font(name='Arial', size=14, bold=True, color=COLOR_GOLD)
        
        ws.merge_cells('B35:J37')
        ws['B35'].value = '=REPT("█", INT(Z1/5)) & "  " & Z1 & " PONTOS"'
        ws['B35'].font = Font(name='Consolas', size=26, bold=True, color=COLOR_ORANGE)
        ws['B35'].alignment = Alignment(horizontal='center', vertical='center')

        # 7. CARD DNA (B41:J48)
        for r in range(41, 49):
            for c in range(2, 11):
                ws.cell(row=r, column=c).fill = fill_card
        
        ws['B41'].value = "🧠 ANÁLISE TÉCNICA IA (DNA)"
        ws['B41'].font = Font(name='Arial', size=14, bold=True, color=COLOR_GOLD)
        
        ws['B43'].value = "IMÓVEL:"
        ws['C43'].value = "---"
        
        dna_labels = [("B45", "Jurídica:", "C45"), ("D45", "Técnica:", "E45"), ("F45", "Geral:", "G45")]
        for lb, t, v in dna_labels:
            ws[lb].value = t
            ws[lb].font = Font(name='Arial', size=10, color=COLOR_GRAY)
            ws[v].font = Font(name='Arial', size=12, bold=True, color=COLOR_GOLD)

        # 8. BOTÃO WHATSAPP (B51:J54)
        ws.merge_cells('B51:J54')
        ws['B51'].value = "📲 Continuar Atendimento via WhatsApp"
        ws['B51'].font = Font(name='Arial', size=18, bold=True, color=COLOR_WHITE)
        ws['B51'].fill = PatternFill(start_color=COLOR_ORANGE, end_color=COLOR_ORANGE, fill_type="solid")
        ws['B51'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B51'].hyperlink = "https://wa.me/5511960364355"

        # 9. TABELA DE FLUXO 36 MESES (A60:J100)
        ws.merge_cells('B60:J60')
        ws['B60'].value = "📅 CRONOGRAMA DE INVESTIMENTO PROGRESSIVO"
        ws['B60'].font = Font(name='Arial', size=14, bold=True, color=COLOR_GOLD)
        ws['B60'].fill = PatternFill(start_color=COLOR_CARD, end_color=COLOR_CARD, fill_type="solid")

        header_fluxo = ["MÊS", "DATA", "CAPITAL", "13º SALÁRIO", "TOTAL", "STATUS"]
        for i, h in enumerate(header_fluxo):
            cell = ws.cell(row=62, column=i+2)
            cell.value = h
            cell.font = Font(name='Arial', size=11, bold=True, color=COLOR_GOLD)
            cell.border = Border(bottom=Side(style='thin', color=COLOR_GOLD))

        # Fórmulas de Fluxo (Google Sheets compatible)
        for i in range(36):
            r = 63 + i
            ws.cell(row=r, column=2, value=i+1).font = Font(color=COLOR_WHITE)
            ws.cell(row=r, column=3, value=f"=EDATE(TODAY(),{i})").font = Font(color=COLOR_WHITE)
            ws.cell(row=r, column=4, value="=$Z$10").font = Font(color=COLOR_WHITE) # Mensalidade Injetada
            ws.cell(row=r, column=5, value=f"=IF(MONTH(C{r})=12, $C$16*0.5, 0)").font = Font(color=COLOR_WHITE)
            ws.cell(row=r, column=6, value=f"=D{r}+E{r}").font = Font(color=COLOR_WHITE)
            ws.cell(row=r, column=7, value="⏳ EM CURSO").font = Font(color=COLOR_GRAY)

        # 10. ÁREA DE LÓGICA OCULTA
        ws['Z1'] = 85 # Score
        ws['Z10'] = 1500 # Mensalidade fixa
        
        ws.freeze_panes = "A9"

        wb.save(path_output)
        print(f"Template v7 Elite Finalizado: {path_output}")

    except Exception as e:
        print(f"Erro Crítico v7: {e}")

if __name__ == "__main__":
    gerar_elite_v7_completo("c:/Users/Marcos.PC_M1/Documents/site_mt/assets/docs/template_mt_parceiros_v7_elite.xlsx")
