import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

def audit_v4(path):
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
        report = []
        report.append(f"### Auditoria Técnica: Template MT Parceiros v4")
        report.append(f"**Abas:** {wb.sheetnames}")
        
        for name in wb.sheetnames:
            ws = wb[name]
            report.append(f"\n--- Aba: {name} ---")
            report.append(f"Dimensões: {ws.dimensions}")
            
            # Amostra de células e fórmulas
            formulas = []
            values = []
            for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=20):
                for cell in row:
                    if cell.value:
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            formulas.append(f"{cell.coordinate}: {cell.value}")
                        else:
                            values.append(f"{cell.coordinate}: {cell.value}")
            
            report.append(f"#### Valores Encontrados (Top 20):")
            for v in values[:20]:
                report.append(f" - {v}")
                
            report.append(f"#### Fórmulas Encontradas (Top 20):")
            if formulas:
                for f in formulas[:20]:
                    report.append(f" - {f}")
            else:
                report.append(" - Nenhuma fórmula detectada nesta aba.")

        with open("report_audit_v4.txt", "w", encoding="utf-8") as f:
            f.write("\n".join(report))
        print("Auditoria v4 concluída.")

    except Exception as e:
        print(f"Erro ao auditar v4: {e}")

if __name__ == "__main__":
    audit_v4("c:/Users/Marcos.PC_M1/Documents/site_mt/assets/docs/template_mt_parceiros_v4.xlsx")
