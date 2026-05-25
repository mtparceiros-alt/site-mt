import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def gerar_supreme_v9_corrigido(path_output):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dossiê Supremo MT"
        ws.sheet_view.showGridLines = False

        # --- PALETA GROK SUPREME ---
        COLOR_BG = "0a0f1c"       
        COLOR_CARD = "121a2b"     
        COLOR_ORANGE = "f35525"   
        COLOR_GOLD = "d4af37"     
        COLOR_WHITE = "ffffff"
        COLOR_GRAY = "b0b8c9"
        COLOR_PURPLE = "8a2be2"
        COLOR_BLUE = "1e90ff"
        COLOR_GREEN = "32cd32"

        # 1. Pintar fundo total
        fill_bg = PatternFill(start_color=COLOR_BG, end_color=COLOR_BG, fill_type="solid")
        for row in range(1, 401):
            for col in range(1, 40):
                ws.cell(row=row, column=col).fill = fill_bg

        # 2. GRID DE ESTABILIDADE
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 6   
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 20

        # 3. HEADER SUPREME
        ws.merge_cells('B2:J5')
        header = ws['B2']
        header.value = "MT PARCEIROS • DOSSIÊ SUPREMO DE VIABILIDADE 2026"
        header.font = Font(name='Arial', size=26, bold=True, color=COLOR_GOLD)
        header.alignment = Alignment(horizontal='center', vertical='center')
        
        ws['B6'].value = "RELATÓRIO TÉCNICO OFICIAL DE CRÉDITO E PLANEJAMENTO FINANCEIRO"
        ws['B6'].font = Font(name='Arial', size=11, bold=True, color=COLOR_GRAY)

        # 4. RESTAURAÇÃO: PERFIL DO CLIENTE (6 Campos - B10:E25)
        fill_card = PatternFill(start_color=COLOR_CARD, end_color=COLOR_CARD, fill_type="solid")
        for r in range(10, 26):
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill_card
        
        ws['B10'].value = "👤 PERFIL DO CLIENTE"
        ws['B10'].font = Font(name='Arial', size=14, bold=True, color=COLOR_GOLD)
        
        lbls = [("B12", "Nome Completo:", "C12"), ("B14", "Renda Bruta (R$):", "C14"),
                ("B15", "Entrada em Mãos:", "C15"), ("B16", "Saldo FGTS:", "C16"),
                ("B17", "Carteira Assinada?", "C17"), ("B18", "Dívidas/Comprom.:", "C18")]
        for lb, t, v in lbls:
            ws[lb].value = t
            ws[lb].font = Font(name='Arial', size=11, color=COLOR_WHITE)
            ws[v].font = Font(name='Arial', size=12, bold=True, color=COLOR_WHITE)
            ws[v].border = Border(bottom=Side(style='thin', color=COLOR_GOLD))

        # 5. RESTAURAÇÃO: LAUDO TÉCNICO (3 Campos + Poder Gigate - G10:J25)
        for r in range(10, 26):
            for c in range(7, 11):
                ws.cell(row=r, column=c).fill = fill_card
        
        ws['G10'].value = "📊 LAUDO TÉCNICO MCMV"
        ws['G10'].font = Font(name='Arial', size=14, bold=True, color=COLOR_GOLD)
        
        ws['G12'].value = "ENQUADRAMENTO:"
        ws['I12'].font = Font(name='Arial', size=12, bold=True, color=COLOR_WHITE)
        
        ws['G14'].value = "SUBSÍDIO LIBERADO:"
        ws['I14'].font = Font(name='Arial', size=14, bold=True, color=COLOR_ORANGE)
        
        ws['G17'].value = "PODER DE COMPRA:"
        ws.merge_cells('G18:J22')
        ws['G18'].font = Font(name='Arial', size=36, bold=True, color=COLOR_ORANGE)
        ws['G18'].alignment = Alignment(horizontal='center', vertical='center')

        # 6. MÓDULO 3: PLANO DE FLUXO ORGANIZADO (Início Linha 28)
        ws['B28'].value = "1. Plano de Fluxo Organizado"
        ws['B28'].font = Font(name='Arial', size=18, bold=True, color=COLOR_WHITE)
        
        def criar_mini_card_v91(ws, start_row, label, subtext, border_color):
            border_side = Side(style='thick', color=border_color)
            for r in range(start_row, start_row+5):
                for c in range(2, 11):
                    ws.cell(row=r, column=c).fill = fill_card
                    if c == 2: ws.cell(row=r, column=c).border = Border(left=border_side)
            ws.cell(row=start_row, column=3).value = label.upper()
            ws.cell(row=start_row, column=3).font = Font(name='Arial', size=10, bold=True, color=COLOR_GRAY)
            ws.merge_cells(start_row=start_row+1, start_column=3, end_row=start_row+2, end_column=9)
            val_cell = ws.cell(row=start_row+1, column=3)
            val_cell.font = Font(name='Arial', size=24, bold=True, color=COLOR_WHITE)
            ws.cell(row=start_row+3, column=3).value = subtext
            ws.cell(row=start_row+3, column=3).font = Font(name='Arial', size=11, color=COLOR_GRAY)
            return val_cell

        criar_mini_card_v91(ws, 31, "ENQUADRAMENTO (SP 2026)", "✓ Benefício MCMV", COLOR_PURPLE) # C32
        criar_mini_card_v91(ws, 38, "ENTRADA PARCELADA (CONSTRUTORA)", "📅 Período: 36 meses", COLOR_ORANGE) # C39
        criar_mini_card_v91(ws, 45, "EVOLUÇÃO DE OBRA (MÉDIA)", "⚒ Conforme a obra avança", COLOR_BLUE) # C46
        criar_mini_card_v91(ws, 52, "PRESTAÇÃO MENSAL (PÓS-CHAVES)", "🏠 Financiamento definitivo", COLOR_GREEN) # C53

        # 7. Restante do Template (Fluxo e WhatsApp)
        ws.merge_cells('B61:J61')
        ws['B61'].value = "📅 CRONOGRAMA TÉCNICO DE INVESTIMENTO (DETALHADO)"
        ws['B61'].font = Font(name='Arial', size=14, bold=True, color=COLOR_GOLD)
        ws['B61'].fill = fill_card
        
        ws.merge_cells('B160:J163')
        ws['B160'].value = "📲 CONTINUAR ATENDIMENTO EXCLUSIVO VIA WHATSAPP"
        ws['B160'].font = Font(name='Arial', size=18, bold=True, color=COLOR_WHITE)
        ws['B160'].fill = PatternFill(start_color=COLOR_ORANGE, end_color=COLOR_ORANGE, fill_type="solid")
        ws['B160'].alignment = Alignment(horizontal='center', vertical='center')

        ws['Z1'] = 85 
        ws['Z10'] = 1500 
        ws.freeze_panes = "A9"
        wb.save(path_output)

    except Exception as e: print(f"Erro: {e}")

if __name__ == "__main__":
    gerar_supreme_v9_corrigido("c:/Users/Marcos.PC_M1/Documents/site_mt/assets/docs/template_mt_parceiros_v9_supreme.xlsx")
