import openpyxl

def audit_v5(path):
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
        report = []
        report.append(f"### Auditoria Técnica: Template MT Parceiros v5 (Charts Version)")
        report.append(f"**Abas:** {wb.sheetnames}")
        
        for name in wb.sheetnames:
            ws = wb[name]
            report.append(f"\n--- Aba: {name} ---")
            report.append(f"Dimensões: {ws.dimensions}")
            
            # Checar Gráficos
            charts = ws._charts if hasattr(ws, '_charts') else []
            report.append(f"#### Gráficos Detectados: {len(charts)}")
            for i, chart in enumerate(charts):
                report.append(f" - Gráfico {i+1}: {type(chart).__name__}")
            
            # Amostra de células e fórmulas (focando em mudanças de mapeamento)
            formulas = []
            values = []
            for row in ws.iter_rows(min_row=1, max_row=60, min_col=1, max_col=15):
                for cell in row:
                    if cell.value:
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            formulas.append(f"{cell.coordinate}: {cell.value}")
                        else:
                            values.append(f"{cell.coordinate}: {cell.value}")
            
            report.append(f"#### Valores e Lógica (Top 30):")
            for v in values[:30]:
                report.append(f" - {v}")
                
            report.append(f"#### Fórmulas e Conexões (Top 20):")
            if formulas:
                for f in formulas[:20]:
                    report.append(f" - {f}")

        with open("report_audit_v5.txt", "w", encoding="utf-8") as f:
            f.write("\n".join(report))
        print("Auditoria v5 concluída.")

    except Exception as e:
        print(f"Erro ao auditar v5: {e}")

if __name__ == "__main__":
    audit_v5("c:/Users/Marcos.PC_M1/Documents/site_mt/assets/docs/template_mt_parceiros_v5_charts.xlsx")
