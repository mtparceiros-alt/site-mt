import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def gerar_supreme_v8(path_output):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dossiê Supremo MT"
        ws.sheet_view.showGridLines = False

        # --- PALETA GROK SUPREME ---
        COLOR_BG = "0a0f1c"       # Azul Noite
        COLOR_CARD = "121a2b"     # Card Visual
        COLOR_ORANGE = "f35525"   # Laranja MT
        COLOR_GOLD = "d4af37"     # Ouro Premium
        COLOR_WHITE = "ffffff"
        COLOR_GRAY = "b0b8c9"

        # 1. Pintar fundo total (A1:AZ400)
        fill_bg = PatternFill(start_color=COLOR_BG, end_color=COLOR_BG, fill_type="solid")
        for row in range(1, 401):
            for col in range(1, 53):
                ws.cell(row=row, column=col).fill = fill_bg

        # 2. GRID DE ESTABILIDADE (Margins & Gutters)
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 6   # Gutter Central
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 20
        ws.column_dimensions['K'].width = 3

        # 3. HEADER SUPREME (B2:J8)
        ws.merge_cells('B2:J5')
        header = ws['B2']
        header.value = "MT PARCEIROS • DOSSIÊ SUPREMO DE VIABILIDADE 2026"
        header.font = Font(name='Arial', size=26, bold=True, color=COLOR_GOLD)
        header.alignment = Alignment(horizontal='center', vertical='center')
        
        ws['B6'].value = "RELATÓRIO TÉCNICO OFICIAL DE CRÉDITO E PLANEJAMENTO FINANCEIRO"
        ws['B6'].font = Font(name='Arial', size=11, bold=True, color=COLOR_GRAY)
        
        ws['I6'].value = "SIMULAÇÃO:"
        ws['I6'].font = Font(name='Arial', size=11, color=COLOR_GRAY)
        ws['J6'].value = "=TEXT(TODAY(), \"dd/mm/yyyy\")"
        ws['J6'].font = Font(name='Arial', size=11, bold=True, color=COLOR_WHITE)

        # 4. MODULO 1: IDENTIFICAÇÃO (B10:E25)
        fill_card = PatternFill(start_color=COLOR_CARD, end_color=COLOR_CARD, fill_type="solid")
        for r in range(10, 26):
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill_card
        
        ws['B10'].value = "👤 PERFIL DO CLIENTE"
        ws['B10'].font = Font(name='Arial', size=14, bold=True, color=COLOR_GOLD)

        lbls_input = [("B12", "Nome Completo:", "C12"), ("B14", "Renda Bruta (R$):", "C14"),
                      ("B15", "Entrada em Mãos:", "C15"), ("B16", "Saldo FGTS:", "C16"),
                      ("B17", "Carteira Assinada?", "C17"), ("B18", "Dívidas/Comprom.:", "C18")]
        for lb, t, v in lbls_input:
            ws[lb].value = t
            ws[lb].font = Font(name='Arial', size=11, color=COLOR_WHITE)
            ws[v].font = Font(name='Arial', size=12, bold=True, color=COLOR_WHITE)
            ws[v].border = Border(bottom=Side(style='thin', color=COLOR_GOLD))

        # 5. MODULO 2: LAUDO DE CRÉDITO (G10:J25)
        for r in range(10, 26):
            for c in range(7, 11):
                ws.cell(row=r, column=c).fill = fill_card
        
        ws['G10'].value = "📊 LAUDO TÉCNICO MCMV"
        ws['G10'].font = Font(name='Arial', size=14, bold=True, color=COLOR_GOLD)
        
        ws['G12'].value = "ENQUADRAMENTO:"
        ws.merge_cells('H12:J12')
        ws['H12'].value = "=IF(C14<=2850, \"FAIXA 1\", IF(C14<=4700, \"FAIXA 2\", \"FAIXA 3\"))"
        ws['H12'].font = Font(name='Arial', size=12, bold=True, color=COLOR_WHITE)

        ws['G14'].value = "SUBSÍDIO LIBERADO:"
        ws['J14'].value = "=IF(C14<=2850, 55000, IF(C14<=4700, 35000, 0))"
        ws['J14'].font = Font(name='Arial', size=14, bold=True, color=COLOR_ORANGE)

        ws['G17'].value = "PODER DE COMPRA:"
        ws.merge_cells('G18:J21')
        ws['G18'].value = "=ROUND(MAX(0,(C14*0.3 - C18)*142) + J14 + C15 + C16, 0)"
        ws['G18'].font = Font(name='Arial', size=36, bold=True, color=COLOR_ORANGE)
        ws['G18'].alignment = Alignment(horizontal='center', vertical='center')

        # 6. MODULO 3: SAÚDE FINANCEIRA (B27:J38)
        for r in range(27, 39):
            for c in range(2, 11):
                ws.cell(row=r, column=c).fill = fill_card
        
        ws['B27'].value = "🏥 ANÁLISE DE SAÚDE E SCORE"
        ws['B27'].font = Font(name='Arial', size=14, bold=True, color=COLOR_GOLD)

        ws['B29'].value = "STATUS DE MARGEM:"
        ws.merge_cells('C29:J29')
        ws['C29'].value = "=IF(C18/C14 > 0.3, \"🔴 RISCO: Suas despesas excedem 30% da renda.\", \"🟢 EXCELENTE: Margem saudável de aprovação.\")"
        ws['C29'].font = Font(name='Arial', size=12, bold=True, color=COLOR_WHITE)

        # Barra de Score (REPT)
        ws.merge_cells('B32:J34')
        ws['B32'].value = '=REPT("█", INT(Z1/5)) & "  " & Z1 & " PONTOS"'
        ws['B32'].font = Font(name='Consolas', size=26, bold=True, color=COLOR_ORANGE)
        ws['B32'].alignment = Alignment(horizontal='center', vertical='center')

        # 7. MODULO 4: FLUXO DE CAIXA SUPREME (A41:J100)
        ws.merge_cells('B41:J41')
        ws['B41'].value = "📅 CRONOGRAMA TÉCNICO DE INVESTIMENTO (36 MESES)"
        ws['B41'].font = Font(name='Arial', size=14, bold=True, color=COLOR_GOLD)
        ws['B41'].fill = fill_card

        headers_fluxo = ["MÊS", "DATA", "MENSALIDADE", "13º SALÁRIO", "FGTS/EXTRA", "TOTAL MÊS", "ACUMULADO"]
        for i, h in enumerate(headers_fluxo):
            cell = ws.cell(row=43, column=i+2)
            cell.value = h
            cell.font = Font(name='Arial', size=11, bold=True, color=COLOR_GOLD)
            cell.border = Border(bottom=Side(style='thin', color=COLOR_GOLD))

        for i in range(36):
            rw = 44 + i
            ws.cell(row=rw, column=2, value=i+1).font = Font(color=COLOR_WHITE)
            ws.cell(row=rw, column=3, value=f"=EDATE(TODAY(),{i})").font = Font(color=COLOR_WHITE)
            ws.cell(row=rw, column=4, value="=$Z$10").font = Font(color=COLOR_WHITE) # Mensalidade JS
            ws.cell(row=rw, column=5, value=f"=IF(MONTH(C{rw})=12, $C$14*0.5, 0)").font = Font(color=COLOR_WHITE)
            ws.cell(row=rw, column=6, value=0).font = Font(color=COLOR_WHITE) # Extra
            ws.cell(row=rw, column=7, value=f"=D{rw}+E{rw}+F{rw}").font = Font(color=COLOR_WHITE)
            if i == 0:
                ws.cell(row=rw, column=8, value=f"=G{rw}").font = Font(color=COLOR_GOLD)
            else:
                ws.cell(row=rw, column=8, value=f"=H{rw-1}+G{rw}").font = Font(color=COLOR_GOLD)

        # 8. MODULO 5: DOSSIÊ DNA & BI (B110:J125)
        for r in range(110, 126):
            for c in range(2, 11):
                ws.cell(row=r, column=c).fill = fill_card

        ws['B110'].value = "🧠 INTELIGÊNCIA DE AUTORIDADE (MT INSIGHT)"
        ws['B110'].font = Font(name='Arial', size=14, bold=True, color=COLOR_GOLD)
        
        ws['B112'].value = "IMÓVEL SELECIONADO:"
        ws['C112'].value = "---"
        ws['C112'].font = Font(name='Arial', size=12, bold=True, color=COLOR_GOLD)

        ws['B114'].value = "VIABILIDADE JURÍDICA:"
        ws['B115'].value = "VISTORIA TÉCNICA:"
        ws['B116'].value = "NOTA BI/CRM:"
        for p in ['B114', 'B115', 'B116']: ws[p].font = Font(name='Arial', size=10, color=COLOR_GRAY)
        
        # 9. BOTÃO WHATSAPP SUPREME (B130:J133)
        ws.merge_cells('B130:J133')
        ws['B130'].value = "📲 CONTINUAR ATENDIMENTO EXCLUSIVO VIA WHATSAPP"
        ws['B130'].font = Font(name='Arial', size=18, bold=True, color=COLOR_WHITE)
        ws['B130'].fill = PatternFill(start_color=COLOR_ORANGE, end_color=COLOR_ORANGE, fill_type="solid")
        ws['B130'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B130'].hyperlink = "https://wa.me/5511960364355"

        # 10. ÁREA OCULTA
        ws['Z1'] = 85 # Score
        ws['Z10'] = 1500 # Mensalidade
        
        ws.freeze_panes = "A9"
        wb.save(path_output)
        print(f"Sucesso: {path_output}")

    except Exception as e:
        print(f"Erro: {e}")

if __name__ == "__main__":
    gerar_supreme_v8("c:/Users/Marcos.PC_M1/Documents/site_mt/assets/docs/template_mt_parceiros_v8_supreme.xlsx")
