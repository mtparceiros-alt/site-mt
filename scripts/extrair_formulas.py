import openpyxl

def detail_all_formulas(path):
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
        report = []
        for name in wb.sheetnames:
            ws = wb[name]
            report.append(f"\n===== ABA: {name} =====")
            formulas = []
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formulas.append(f"Célula {cell.coordinate}: {cell.value}")
            
            if formulas:
                report.extend(formulas)
            else:
                report.append("Nenhuma fórmula encontrada nesta aba.")
        
        with open("log_formulas_excel.txt", "w", encoding="utf-8") as f:
            f.write("\n".join(report))
        print("Log de fórmulas gerado com sucesso.")

    except Exception as e:
        print(f"Erro: {e}")

if __name__ == "__main__":
    detail_all_formulas("c:/Users/Marcos.PC_M1/Documents/site_mt/assets/docs/template_mt_parceiros.xlsx")
