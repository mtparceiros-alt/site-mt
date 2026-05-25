import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
import os

def gerar_super_dossie_v3(path_original, path_novo):
    try:
        # Criar novo workbook com uma única aba
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dossiê MT 2026"
        ws.sheet_view.showGridLines = False

        # --- ESTILOS PADRÃO (Google Sheets Compatible) ---
        fill_bg = PatternFill(start_color="121212", end_color="121212", fill_type="solid")
        fill_card = PatternFill(start_color="1E1E1E", end_color="1E1E1E", fill_type="solid")
        font_gold = Font(name='Arial', size=12, bold=True, color="C5A030")
        font_white = Font(name='Arial', size=10, color="FFFFFF")
        font_header = Font(name='Arial', size=18, bold=True, color="C5A030")
        border_thin = Border(left=Side(style='thin', color='333333'), 
                             right=Side(style='thin', color='333333'), 
                             top=Side(style='thin', color='333333'), 
                             bottom=Side(style='thin', color='333333'))

        # Pintar fundo de grafite (A1:AZ100)
        for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=50):
            for cell in row:
                cell.fill = fill_bg

        # --- CABEÇALHO (A1:J5) ---
        ws.merge_cells('A1:J5')
        header = ws['A1']
        header.value = "MT PARCEIROS | DOSSIÊ INTELIGENTE 2026"
        header.font = font_header
        header.alignment = Alignment(horizontal='center', vertical='center')

        # --- BLOCO 1: PERFIL FINANCEIRO (A7:E15) ---
        ws.merge_cells('A7:E7')
        box1_title = ws['A7']
        box1_title.value = "👤 PERFIL DO CLIENTE"
        box1_title.font = font_gold
        box1_title.fill = fill_card

        labels = [("A8", "NOME:"), ("A9", "RENDA BRUTA:"), ("A10", "FGTS:"), 
                  ("A11", "ENTRADA:"), ("A12", "DÍVIDAS:"), ("A13", "VÍNCULO:"), 
                  ("A14", "DEPENDENTES:")]
        for pos, text in labels:
            ws[pos] = text
            ws[pos].font = font_white
            # Célula de valor (E):
            val_pos = pos.replace("A", "E")
            ws[val_pos].fill = fill_card
            ws[val_pos].font = font_white
            ws[val_pos].border = border_thin

        # --- BLOCO 2: PAINEL POWER BI (G7:J15) ---
        ws.merge_cells('G7:J7')
        box2_title = ws['G7']
        box2_title.value = "📊 ANÁLISE DE CRÉDITO IA"
        box2_title.font = font_gold
        box2_title.fill = fill_card

        ws['G8'] = "MT SCORE:"
        ws['G8'].font = font_white
        ws.merge_cells('H8:J8')
        ws['H8'] = '=REPT("█", INT(E28/10)) & " " & E28 & "%"' # Simulação de Barra de Progresso
        ws['H8'].font = font_gold

        ws['G10'] = "STATUS:"
        ws['G10'].font = font_white
        ws['J10'].fill = fill_card
        ws['J10'].font = Font(bold=True, color="00FF00")

        ws['G12'] = "PODER TOTAL:"
        ws['G12'].font = font_white
        ws['J12'].fill = fill_card

        # --- BLOCO 3: DOSSIÊ DNA (A17:J21) ---
        ws.merge_cells('A17:J17')
        box3_title = ws['A17']
        box3_title.value = "🛡️ DOSSIÊ DNA (AUTORIDADE TÉCNICA)"
        box3_title.font = font_gold
        box3_title.fill = fill_card

        ws['A18'] = "IMÓVEL:"
        ws.merge_cells('B18:J18')
        ws['A19'] = "Auditoria Jurídica:"
        ws['D19'] = "Vistoria Técnica:"
        ws['G19'] = "Viabilidade Finan:"
        for p in ['A19', 'D19', 'G19']: ws[p].font = font_white

        # --- BLOCO 4: FLUXO 36 MESES (A23:J60) ---
        ws.merge_cells('A23:J23')
        ws['A23'] = "📅 CRONOGRAMA DE INVESTIMENTO (MÉDULA FINANCEIRA)"
        ws['A23'].font = font_gold
        ws['A23'].fill = fill_card

        cols = ["MÊS", "DATA", "CAPITAL", "13º SALÁRIO", "TOTAL MENSAL", "STATUS", "CONQUISTA (%)"]
        for i, col in enumerate(cols):
            cell = ws.cell(row=24, column=i+1)
            cell.value = col
            cell.font = font_gold
            cell.border = border_thin

        # Adicionar 36 linhas de fluxo com fórmulas
        start_row = 25
        for i in range(36):
            curr_row = start_row + i
            ws.cell(row=curr_row, column=1, value=i+1).font = font_white
            # Data: =DATA.MÊS(HOJE(); i) - Simplificado para Google Sheets
            ws.cell(row=curr_row, column=2, value=f"=EDATE(TODAY(), {i})").font = font_white
            # Valor fixo da parcela (obtido do input)
            ws.cell(row=curr_row, column=3, value="=$E$45").font = font_white
            # Lógica 13º: =SE(MÊS(B{curr_row})=12; $E$9*0.5; 0)
            ws.cell(row=curr_row, column=4, value=f"=IF(MONTH(B{curr_row})=12, $E$9*0.5, 0)").font = font_white
            # Total: =C{curr_row} + D{curr_row}
            ws.cell(row=curr_row, column=5, value=f"=C{curr_row}+D{curr_row}").font = font_white

        # --- ÁREA DE LÓGICA OCULTA (AA1:AH50) ---
        ws['AA1'] = "MOTOR DE INTELIGÊNCIA ( NÃO REMOVER )"
        ws['AA1'].font = Font(color="333333")
        
        # Inserindo Tabelas de Referência MCMV
        ws['AA3'] = "Teto MCMV:"
        ws['AB3'] = 350000
        ws['AA4'] = "Taxa MCMV 1:"
        ws['AB4'] = 0.045

        # Célula de Cálculo de Score (E28) que o KPI lê
        ws['E28'] = "=MIN(100, 40 + (E9/1000) + (E10/5000))" # Exemplo de Score dinâmico

        # Salvar
        wb.save(path_novo)
        print(f"Template v3 gerado: {path_novo}")

    except Exception as e:
        print(f"Erro ao gerar template v3: {e}")

if __name__ == "__main__":
    gerar_super_dossie_v3(
        "c:/Users/Marcos.PC_M1/Documents/site_mt/assets/docs/template_mt_parceiros.xlsx",
        "c:/Users/Marcos.PC_M1/Documents/site_mt/assets/docs/template_mt_parceiros_v3.xlsx"
    )
