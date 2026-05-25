import openpyxl

def audit_template(path):
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
        print(f"Auditando: {path}")
        print(f"Abas encontradas: {wb.sheetnames}")
        
        for name in wb.sheetnames:
            ws = wb[name]
            print(f"\n--- Aba: {name} ---")
            print(f"Dimensões: {ws.dimensions}")
            
            # Checar células com fórmulas
            formulas = []
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formulas.append(f"{cell.coordinate}: {cell.value}")
            
            print(f"Fórmulas encontradas: {len(formulas)}")
            if formulas:
                for f in formulas[:10]: # Mostrar as 10 primeiras
                    print(f"  {f}")
                    
            # Checar imagens (se houver)
            if hasattr(ws, '_images'):
                print(f"Imagens encontradas: {len(ws._images)}")

    except Exception as e:
        print(f"Erro ao auditar: {e}")

if __name__ == "__main__":
    audit_template("c:/Users/Marcos.PC_M1/Documents/site_mt/assets/docs/template_mt_parceiros.xlsx")
