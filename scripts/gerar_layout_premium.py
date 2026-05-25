import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def aplicar_design_premium(path_input, path_output):
    try:
        wb = openpyxl.load_workbook(path_input, data_only=False)
        ws = wb['MT Parceiros']
        ws.sheet_view.showGridLines = False

        # --- LIMPEZA DE MERGES ORIGINAIS (Para evitar erro de read-only) ---
        # Salvamos uma cópia da lista para evitar erro ao iterar/remover
        merges = list(ws.merged_cells.ranges)
        for m in merges:
            ws.unmerge_cells(str(m))

        # --- PALETA DE CORES ---
        COLOR_BG = "0a0f1c"       
        COLOR_CARD = "121a2b"     
        COLOR_ORANGE = "f35525"   
        COLOR_GOLD = "d4af37"     
        COLOR_WHITE = "ffffff"
        COLOR_GRAY = "cccccc"

        # 1. Pintar fundo total (A1:AZ200)
        fill_bg = PatternFill(start_color=COLOR_BG, end_color=COLOR_BG, fill_type="solid")
        for row in range(1, 201):
            for col in range(1, 53):
                ws.cell(row=row, column=col).fill = fill_bg

        # 2. HEADER PREMIUM (B2:J4)
        ws['B2'].value = "RELATÓRIO INTELIGENTE DE VIABILIDADE IMOBILIÁRIA"
        ws['B2'].font = Font(name='Arial', size=24, bold=True, color=COLOR_GOLD)
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('B2:J4')
        
        ws['B5'].value = "SIMULAÇÃO OFICIAL | MT PARCEIROS"
        ws['B5'].font = Font(name='Arial', size=11, bold=True, color=COLOR_GRAY)
        
        ws['I5'].value = "DATA:"
        ws['I5'].font = Font(name='Arial', size=11, color=COLOR_GRAY)
        ws['J5'].value = "=TODAY()" # Forçar data atual
        ws['J5'].font = Font(name='Arial', size=11, bold=True, color=COLOR_WHITE)

        # 3. BLOCO DE ENTRADAS (INPUTS) - Card Visual (B7:E25)
        fill_card = PatternFill(start_color=COLOR_CARD, end_color=COLOR_CARD, fill_type="solid")
        for r in range(7, 26):
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill_card
        
        # Labels
        input_map = {
            "B7": "📥 PERFIL FINANCEIRO",
            "B9": "NOME DO CLIENTE:",
            "B12": "RENDA BRUTA (R$):",
            "B13": "ENTRADA DINHEIRO (R$):",
            "B14": "SALDO FGTS (R$):",
            "B16": "DESPESAS TOTAIS (R$):"
        }
        for pos, val in input_map.items():
            ws[pos].value = val
            ws[pos].font = Font(name='Arial', size=12, bold=(pos=="B7"), color=(COLOR_GOLD if pos=="B7" else COLOR_WHITE))

        border_gold = Border(left=Side(style='thin', color=COLOR_GOLD), 
                            right=Side(style='thin', color=COLOR_GOLD), 
                            top=Side(style='thin', color=COLOR_GOLD), 
                            bottom=Side(style='thin', color=COLOR_GOLD))
        
        # Destacar os campos de input (C7, C12, C13, C14, C22)
        for i in ['C7', 'C12', 'C13', 'C14', 'C22']:
            ws[i].border = border_gold
            ws[i].font = Font(name='Arial', size=12, bold=True, color=COLOR_WHITE)

        # 4. BLOCO DE RESULTADOS (G7:J25)
        for r in range(7, 26):
            for c in range(7, 11):
                ws.cell(row=r, column=c).fill = fill_card

        ws['G7'].value = "🎯 PARECER TÉCNICO"
        ws['G7'].font = Font(name='Arial', size=14, bold=True, color=COLOR_GOLD)

        ws['G9'].value = "PODER DE COMPRA ESTIMADO:"
        ws['G9'].font = Font(name='Arial', size=11, color=COLOR_WHITE)
        
        ws['G10'].value = "=D27" # Valor que já existe na v5
        ws['G10'].font = Font(name='Arial', size=32, bold=True, color=COLOR_ORANGE)
        ws['G10'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('G10:J12')

        ws['G14'].value = "VALOR FINANCIADO (ESTIMADO):"
        ws['G14'].font = Font(name='Arial', size=11, color=COLOR_WHITE)
        ws['G15'].value = "=IFERROR(D27 - C13 - C14, 0)" 
        ws['G15'].font = Font(name='Arial', size=20, bold=True, color=COLOR_WHITE)

        # 5. BLOCO SCORE / APROVAÇÃO (B27:J32)
        for r in range(27, 33):
            for c in range(2, 11):
                ws.cell(row=r, column=c).fill = fill_card

        ws['B27'].value = "🏥 SCORE DE JORNADA MT"
        ws['B27'].font = Font(name='Arial', size=14, bold=True, color=COLOR_GOLD)
        
        # D43 tem o score do original
        ws['B29'].value = '=REPT("█", INT(D43/5)) & " " & D43 & " PONTOS"'
        ws['B29'].font = Font(name='Consolas', size=24, bold=True, color=COLOR_ORANGE)
        ws['B29'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('B29:J31')

        # 6. ANÁLISE DNA (B34:J43)
        for r in range(34, 44):
            for c in range(2, 11):
                ws.cell(row=r, column=c).fill = fill_card
        
        ws['B34'].value = "🧠 ANÁLISE TÉCNICA IA (DNA)"
        ws['B34'].font = Font(name='Arial', size=14, bold=True, color=COLOR_GOLD)
        
        labels_dna = [("B36", "JURÍDICO:"), ("D36", "TÉCNICO:"), ("F36", "FINANCEIRO:"), ("H36", "VALORIZAÇÃO:")]
        for p, l in labels_dna:
            ws[p].value = l
            ws[p].font = Font(name='Arial', size=10, bold=True, color=COLOR_GRAY)
            val_p = get_column_letter(ws[p].column + 1) + str(ws[p].row)
            ws[val_p].value = "9.5"
            ws[val_p].font = Font(name='Arial', size=11, bold=True, color=COLOR_GOLD)

        # 7. BOTÃO WHATSAPP (B46:J49)
        ws['B46'].value = "📲 Continuar Atendimento via WhatsApp"
        ws['B46'].font = Font(name='Arial', size=18, bold=True, color=COLOR_WHITE)
        ws['B46'].fill = PatternFill(start_color=COLOR_ORANGE, end_color=COLOR_ORANGE, fill_type="solid")
        ws['B46'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('B46:J49')
        ws['B46'].hyperlink = "https://wa.me/5511960364355"

        # 8. Finalização
        ws.freeze_panes = "A6"
        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['J'].width = 20

        wb.save(path_output)
        print(f"Sucesso: {path_output}")

    except Exception as e:
        print(f"Erro Crítico: {e}")

if __name__ == "__main__":
    aplicar_design_premium(
        "c:/Users/Marcos.PC_M1/Documents/site_mt/assets/docs/template_mt_parceiros_v5_charts.xlsx",
        "c:/Users/Marcos.PC_M1/Documents/site_mt/assets/docs/template_mt_parceiros_v6_premium.xlsx"
    )
