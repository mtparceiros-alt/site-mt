# ============================================================
#  sync_cms.py — MT Parceiros CMS Sync v2.0
#
#  Lê a planilha Empreendimentos.xlsx (raiz do projeto) e gera
#  o arquivo empreendimentos.js com TODOS os campos necessários
#  para o site funcionar: cards, mapa, drawer, simulador.
#
#  USO:
#    python scripts/sync_cms.py              → Gera empreendimentos.js
#    python scripts/sync_cms.py --dry-run    → Prévia sem salvar
#
#  NOVIDADES (Mar/2026):
#    - IA de Extensões: Busca automática por .png/.jpg se faltar no Excel.
#    - IA de Preços: Suporte a "milhão/milhões" e correção gramatical automática.
#    - Geocodificação v2.1: Prioridade Invertida (Bairro + Endereço) para evitar
#      erros de CEP e de ruas com nomes duplicados em bairros diferentes.
#    - [Abr/2026] Protocolo Anti-Espaço: Recomenda-se nomes sem espaços (ex: em-breve.png)
#      para garantir compatibilidade máxima em servidores web.
#
#  REQUISITOS:
#    pip install openpyxl
#
#  FLUXO:
#    1. Lê cada linha do Excel (colunas A-J)
#    2. Valida campos obrigatórios (nome, endereço, preço)
#    3. Geocodifica endereço (Cache → Google → Nominatim)
#    4. Verifica se a imagem existe na pasta
#    5. Normaliza o preço para o formato "XXXmil"
#    6. Formata a data de entrega (datetime → "Mês/Ano")
#    7. Gera empreendimentos.js (com backup automático)
#
#  ARQUIVOS CONSUMIDOS:
#    - Empreendimentos.xlsx  (fonte de dados)
#    - coords_cache.json     (cache de coordenadas)
#    - .env                  (opcional: GOOGLE_MAPS_API_KEY)
#
#  ARQUIVOS GERADOS:
#    - empreendimentos.js      (banco de dados do site)
#    - empreendimentos.js.bkp  (backup do anterior)
#    - coords_cache.json       (atualizado com novas coords)
# ============================================================

import json
import time
import sys
import os
import re
import urllib.request
import urllib.parse
from datetime import datetime

# ── Forçar saída UTF-8 no Windows ─────────────────────────────
# O console do Windows usa cp1252 por padrão, que não suporta
# emojis e caracteres especiais. Isso força UTF-8.
if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if sys.stderr and hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

# ── Dependência: openpyxl ─────────────────────────────────────
# Biblioteca para leitura de arquivos Excel (.xlsx).
# Instalação: pip install openpyxl
try:
    import openpyxl
except ImportError:
    print("=" * 55)
    print("  ❌ ERRO: Módulo 'openpyxl' não encontrado.")
    print("  Execute no terminal:")
    print("     pip install openpyxl")
    print("=" * 55)
    sys.exit(1)


# ── Configurações ─────────────────────────────────────────────
# Caminhos relativos à raiz do projeto (site_mt/)
ARQUIVO_XLSX  = "Empreendimentos.xlsx"      # Planilha fonte
ARQUIVO_JS    = "empreendimentos.js"        # Saída consumida pelo site
ARQUIVO_CACHE = "coords_cache.json"         # Cache de geocodificação
PASTA_IMAGENS = "assets/images/empreendimentos"  # Pasta onde ficam as fotos

# Correções Manuais de Coordenadas (GPS validado no Google Maps)
# Use quando o Nominatim retorna posição errada para um endereço.
# Formato: "trecho do endereço" → {lat, lng}
FIXES = {
    "Rua Laguna, 440":                     {"lat": -23.63959, "lng": -46.71574},
    "Avenida Cristo Rei, 34":              {"lat": -23.48938, "lng": -46.72078},
    "Professor Miguel Franchini Neto, 186": {"lat": -23.44355, "lng": -46.72530},
    "Rua Professor Clemente Pastore, 14":  {"lat": -23.51383, "lng": -46.69922},
    "Rua Joao Alfredo, 431":               {"lat": -23.65701, "lng": -46.70044},
    "Avenida Condessa Elisabeth de Robiano, 5748": {"lat": -23.51590, "lng": -46.55550},
}

# Limites geográficos de São Paulo (caixa de validação)
# Se uma coordenada cair fora dessa área, é descartada.
LAT_MIN, LAT_MAX = -23.85, -23.35
LNG_MIN, LNG_MAX = -46.95, -46.35

# ── Índices das Colunas no Excel ──────────────────────────────
# A planilha Empreendimentos.xlsx deve ter estas colunas:
#   A=Nome | B=Bairro | C=Endereço | D=Área | E=Quartos
#   F=Diferenciais | G=Lazer | H=Preço | I=Entrega | J=Imagem
COL_NOME    = 0   # A: Nome do empreendimento
COL_BAIRRO  = 1   # B: Bairro / Região (Ex: "Penha (Zona Leste)")
COL_END     = 2   # C: Endereço completo
COL_AREA    = 3   # D: Metragem (Ex: "De 24m² a 38m²")
COL_QUARTOS = 4   # E: Quantidade de quartos (Ex: "1 a 2")
COL_DIFER   = 5   # F: Diferenciais (Ex: "Opção de vaga e varanda")
COL_LAZER   = 6   # G: Facilidades de lazer
COL_PRECO   = 7   # H: Preço a partir de (Ex: "235mil" ou "R$ 235.000")
COL_ENTREGA = 8   # I: Data de entrega (Ex: "Nov/2028", "Pronto", "Lançamento")
COL_IMAGEM  = 9   # J: Nome do arquivo de imagem (Ex: "Vivaz Penha.jpg")


# ══════════════════════════════════════════════════════════════
#  FUNÇÕES DE CACHE E GEOCODIFICAÇÃO
# ══════════════════════════════════════════════════════════════

def carregar_cache():
    """Carrega o arquivo coords_cache.json com coordenadas já geocodificadas.

    O cache evita consultar o Nominatim/Google repetidamente para 
    endereços que já foram processados em execuções anteriores.

    Returns:
        dict: Dicionário {endereço: {lat, lng}} ou {} se não existir.
    """
    if os.path.exists(ARQUIVO_CACHE):
        try:
            with open(ARQUIVO_CACHE, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError) as e:
            print(f"  ⚠️  Cache corrompido, será recriado. Erro: {e}")
            return {}
    return {}


def salvar_cache(cache):
    """Salva o dicionário de coordenadas no arquivo JSON de cache.

    Args:
        cache: Dicionário {endereço: {lat, lng}} a ser salvo.
    """
    with open(ARQUIVO_CACHE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


def consultar_viacep(cep):
    """Consulta o serviço ViaCEP para obter dados do endereço a partir do CEP.

    Útil para melhorar a precisão da geocodificação, pois o Nominatim
    funciona melhor com o nome oficial da rua (retornado pelo ViaCEP).

    Args:
        cep: String com CEP (aceita "01234-567" ou "01234567").

    Returns:
        dict com {logradouro, bairro, cidade, uf} ou None se falhar.
    """
    if not cep:
        return None
    cep_limpo = re.sub(r"\D", "", str(cep))
    if len(cep_limpo) != 8:
        return None

    try:
        url = f"https://viacep.com.br/ws/{cep_limpo}/json/"
        with urllib.request.urlopen(url, timeout=5) as resp:
            data = json.loads(resp.read().decode())
            if "erro" not in data:
                return {
                    "logradouro": data.get("logradouro", ""),
                    "bairro": data.get("bairro", ""),
                    "cidade": data.get("localidade", ""),
                    "uf": data.get("uf", ""),
                }
    except Exception as e:
        print(f"  ⚠️  ViaCEP falhou para CEP {cep}: {e}")
    return None


def geocodificar_google(endereco, bairro):
    """Geocodifica um endereço usando a API do Google Maps.

    Requer a variável GOOGLE_MAPS_API_KEY no arquivo .env.
    Mais precisa que o Nominatim, mas tem custo por consulta.

    Args:
        endereco: Endereço da rua (Ex: "Av. Itaberaba, 1234").
        bairro: Bairro para refinar a busca.

    Returns:
        Tupla (lat, lng, "Google") ou (None, None, None) se falhar.
    """
    if not GOOGLE_KEY:
        return None, None, None

    addr_clean = f"{endereco}, {bairro}, Sao Paulo, SP, Brazil"
    url = (
        f"https://maps.googleapis.com/maps/api/geocode/json"
        f"?address={urllib.parse.quote(addr_clean)}&key={GOOGLE_KEY}"
    )

    try:
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode())
            if data["status"] == "OK":
                loc = data["results"][0]["geometry"]["location"]
                return loc["lat"], loc["lng"], "Google"
    except Exception as e:
        print(f"  ⚠️  Google Maps falhou: {e}")
    return None, None, None


def geocodificar_nominatim(endereco, bairro, cep_info=None):
    """Geocodifica um endereço usando o Nominatim (OpenStreetMap) — gratuito.

    Primeiro verifica se há uma correção manual (FIXES).
    Depois tenta múltiplas consultas com variações do endereço.
    Respeita o rate-limit de 1 consulta por 1.2 segundos.

    Args:
        endereco: Endereço da rua.
        bairro: Bairro para refinar a busca.
        cep_info: Dados do ViaCEP (opcional, melhora precisão).

    Returns:
        Tupla (lat, lng, tipo) ou (None, None, None) se falhar.
        tipo pode ser: "Fixa" (correção manual), "Real" (Nominatim).
    """
    headers = {"User-Agent": "MTParceirosGeocoder/2.0"}
    addr_clean = endereco.replace("\n", " ").strip()
    
    # 0. Limpeza do Bairro (Remove o que estiver entre parênteses, ex: "(Zona Norte)")
    # Isso evita que o GPS se confunda com descrições extras que não fazem parte
    # do nome oficial do bairro no serviço de mapas.
    bairro_limpo = re.sub(r"\(.*?\)", "", bairro).strip() if bairro else ""

    # 1. Verificar correções manuais (FIXES)
    # Se o endereço estiver em FIXES, retornamos as coordenadas validadas na hora.
    for partial_match, coords in FIXES.items():
        if partial_match.lower() in addr_clean.lower():
            return coords["lat"], coords["lng"], "Fixa"

    # 2. Montar lista de consultas alternativas (DA MAIS ESPECÍFICA PARA A MENOS)
    # A ordem aqui é CRÍTICA. Tentamos primeiro o que for mais difícil de errar.
    consultas = []

    # A) Com CEP (Se houver - Maior Precisão)
    # O CEP é o identificador único mais forte de um endereço.
    if cep_info and cep_info["logradouro"]:
        numero = re.search(r"(\d+)", addr_clean)
        num_str = numero.group(1) if numero else ""
        consultas.append(
            f"{cep_info['logradouro']} {num_str}, {cep_info['cidade']}, {cep_info['uf']}, Brazil"
        )

    # B) Endereço + Bairro LIMPO + São Paulo (Estratégia Principal)
    # Resolve problemas como "Rua Piauí" que existe em Higienópolis e em outros bairros.
    # Ao forçar o Bairro na pesquisa, garantimos que o pino caia no lugar certo.
    if bairro_limpo:
        consultas.append(f"{addr_clean}, {bairro_limpo}, Sao Paulo, Brazil")
    
    # C) Endereço + Bairro ORIGINAL + São Paulo (Caso o parênteses seja importante)
    # Fallback caso a limpeza do bairro tenha removido algo essencial (raro).
    if bairro and bairro != bairro_limpo:
        consultas.append(f"{addr_clean}, {bairro}, Sao Paulo, Brazil")

    # D) Endereço + São Paulo (Último recurso - pode gerar falso positivo)
    # Se chegarmos aqui e houver duas ruas com o mesmo nome em SP, o GPS pode errar.
    consultas.append(f"{addr_clean}, Sao Paulo, SP, Brazil")

    # 3. Tentar cada consulta
    for q in consultas:
        try:
            q_encoded = urllib.parse.quote(q)
            url = (
                f"https://nominatim.openstreetmap.org/search"
                f"?q={q_encoded}&format=json&limit=3"
                f"&bounded=1&viewbox=-46.85,-23.35,-46.35,-23.85"
            )

            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read().decode())
                for result in data:
                    lat = float(result["lat"])
                    lon = float(result["lon"])
                    # Validar que está dentro dos limites de São Paulo
                    if LAT_MIN < lat < LAT_MAX and LNG_MIN < lon < LNG_MAX:
                        return lat, lon, "Real"
            time.sleep(1.2)  # Rate-limit: máximo 1 req/1.2s
        except Exception as e:
            print(f"  ⚠️  Nominatim falhou para '{q}': {e}")

    return None, None, None


# ══════════════════════════════════════════════════════════════
#  FUNÇÕES DE FORMATAÇÃO E VALIDAÇÃO
# ══════════════════════════════════════════════════════════════

def formatar_entrega(valor):
    """Converte a data de entrega do Excel para o formato do site.

    O Excel pode salvar datas como datetime (Ex: 2028-11-01 00:00:00).
    O site espera: "Nov/2028", "Pronto" ou "Lançamento".

    Args:
        valor: Valor da célula do Excel (string ou datetime).

    Returns:
        String formatada (Ex: "Nov/2028").
    """
    if not valor:
        return ""
    val = str(valor).strip()

    # Se for datetime do Excel (Ex: "2028-11-01 00:00:00")
    if "00:00:00" in val:
        try:
            dt = datetime.strptime(val, "%Y-%m-%d %H:%M:%S")
            meses = [
                "Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
                "Jul", "Ago", "Set", "Out", "Nov", "Dez",
            ]
            return f"{meses[dt.month - 1]}/{dt.year}"
        except ValueError as e:
            print(f"  ⚠️  Data não reconhecida '{val}': {e}")
            return val

    return val


def normalizar_preco(valor):
    """Normaliza o preço para o formato do site sem quebrar valores na casa dos milhões.
    """
    if not valor:
        return ""
    val = str(valor).strip().lower()

    if re.match(r"^\d+mil$", val) or "milhão" in val or "milhões" in val:
        # Pega a string original que o usuário digitou
        formatado = str(valor).strip()
        # Correção Gramatical: Se tiver "8milhão" ou "2 milhão" (qualquer número maior que 1) -> Troca para "milhões"
        formatado = re.sub(r"([2-9]|\d{2,})\s*milh[aã]o", r"\1 milhões", formatado, flags=re.IGNORECASE)
        # Padroniza espaçamento visual para ficar elegante
        formatado = formatado.replace("milhões", " milhões").replace("milhão", " milhão")
        formatado = re.sub(r"\s+", " ", formatado) # remove eventual espaço duplo se tiver
        
        return formatado.strip()

    # Extrair apenas os dígitos
    digitos = re.sub(r"[^\d]", "", val)
    if not digitos:
        return str(valor).strip()

    num = int(digitos)
    if num >= 1000000:
        return f"{num // 1000000}milhões"
    if num >= 1000:
        return f"{num // 1000}mil"
    return f"{num}mil"


def verificar_imagem(nome_arquivo):
    """Verifica se a imagem existe na pasta de empreendimentos.

    A pasta é: assets/images/empreendimentos/
    Retorna o caminho relativo completo para uso no empreendimentos.js.
    Agora com IA para auto-completar .png ou .jpg ausentes na planilha.

    Args:
        nome_arquivo: Nome do arquivo (Ex: "Vivaz Penha.jpg" ou "Vivaz Penha").

    Returns:
        String com caminho relativo (Ex: "assets/images/empreendimentos/Vivaz Penha.jpg")
        ou string vazia se a imagem não existir.
    """
    if not nome_arquivo:
        return ""
    nome_limpo = nome_arquivo.strip()
    caminho = os.path.join(PASTA_IMAGENS, nome_limpo)
    
    # 1. Tenta exatamente como escrito no Excel
    if os.path.exists(caminho):
        return f"{PASTA_IMAGENS}/{nome_limpo}"
        
    # 2. Tenta forçar as extensões mais comuns se o usuário esqueceu
    # Prioridade para .jpg e .jpeg (Otimizadas para Web)
    if os.path.exists(caminho + ".jpg"):
        return f"{PASTA_IMAGENS}/{nome_limpo}.jpg"
        
    if os.path.exists(caminho + ".jpeg"):
        return f"{PASTA_IMAGENS}/{nome_limpo}.jpeg"

    if os.path.exists(caminho + ".png"):
        return f"{PASTA_IMAGENS}/{nome_limpo}.png"
        
    return ""



# ══════════════════════════════════════════════════════════════
#  FUNÇÃO PRINCIPAL
# ══════════════════════════════════════════════════════════════

def main():
    """Ponto de entrada do script.

    Lê a planilha Empreendimentos.xlsx, processa cada empreendimento
    (geocodificação, imagem, preço, entrega) e gera o empreendimentos.js.

    Flags:
        --dry-run : Mostra prévia no terminal sem salvar o arquivo.
    """
    # ── Flags de linha de comando ─────────────────────────────
    dry_run = "--dry-run" in sys.argv

    print("=" * 55)
    print("  MT Parceiros — CMS Sync v2.0")
    print("  Excel → empreendimentos.js")
    if dry_run:
        print("  ⚡ MODO DRY-RUN (nenhum arquivo será alterado)")
    print("=" * 55)

    # ── Verificar que a planilha existe ───────────────────────
    if not os.path.exists(ARQUIVO_XLSX):
        print(f"\n❌ Arquivo '{ARQUIVO_XLSX}' não encontrado na raiz do projeto.")
        print(f"   Coloque a planilha em: {os.path.abspath(ARQUIVO_XLSX)}")
        sys.exit(1)

    # ── Carregar cache e planilha ─────────────────────────────
    cache = carregar_cache()
    wb = openpyxl.load_workbook(ARQUIVO_XLSX)
    ws = wb.active
    empreendimentos = []
    erros_geocod = 0
    imagens_faltantes = []

    linhas = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
    total = sum(1 for r in linhas if r[COL_NOME])
    print(f"\n📋 Processando {total} empreendimento(s)...\n")

    # ── Processar cada linha ──────────────────────────────────
    contador = 0
    for row in linhas:
        nome = str(row[COL_NOME] or "").strip()
        if not nome:
            continue
        contador += 1

        endereco = str(row[COL_END] or "").strip().replace("\n", " ")
        bairro = str(row[COL_BAIRRO] or "").strip()
        preco_raw = str(row[COL_PRECO] or "").strip()

        # ── Validação de campos obrigatórios ──────────────────
        if not endereco:
            print(f"  ⚠️  [{contador}/{total}] {nome} — SEM ENDEREÇO, pulando.")
            erros_geocod += 1
            continue
        if not preco_raw:
            print(f"  ⚠️  [{contador}/{total}] {nome} — SEM PREÇO, pulando.")
            continue

        print(f"🔍 [{contador}/{total}] {nome}")

        # ── CEP (se houver no endereço) ───────────────────────
        match_cep = re.search(r"(\d{5}-?\d{3})", f"{endereco} {bairro}")
        cep_encontrado = match_cep.group(1) if match_cep else None
        cep_info = consultar_viacep(cep_encontrado) if cep_encontrado else None
        if cep_info:
            print(f"  📮 (ViaCEP) {cep_info['logradouro']} — {cep_info['bairro']}")

        # ── Geocodificação ────────────────────────────────────
        lat, lng = None, None
        if endereco in cache:
            lat = cache[endereco]["lat"]
            lng = cache[endereco]["lng"]
            print(f"  🧊 (Cache) {lat:.5f}, {lng:.5f}")
        else:
            # Tentar Google Maps (se houver API key no .env)
            if GOOGLE_KEY:
                lat, lng, tipo = geocodificar_google(endereco, bairro)

            # Fallback: Nominatim (gratuito, OpenStreetMap)
            if not lat:
                lat, lng, tipo = geocodificar_nominatim(endereco, bairro, cep_info)

            if lat:
                print(f"  📍 ({tipo}) {lat:.5f}, {lng:.5f}")
                cache[endereco] = {"lat": lat, "lng": lng}
                if not dry_run:
                    salvar_cache(cache)
            else:
                print(f"  ❌ Geocodificação falhou para: {endereco}")
                erros_geocod += 1
                continue

        # ── Verificar imagem ──────────────────────────────────
        imagem_nome = str(row[COL_IMAGEM] or "").strip()
        imagem_path = verificar_imagem(imagem_nome)
        if imagem_path:
            print(f"  🖼️  {imagem_path}")
        elif imagem_nome:
            print(f"  ⚠️  Imagem não encontrada: {PASTA_IMAGENS}/{imagem_nome}")
            imagens_faltantes.append(imagem_nome)

        # ── Normalizar preço ──────────────────────────────────
        preco_final = normalizar_preco(preco_raw)

        # ── Montar objeto ─────────────────────────────────────
        empreendimentos.append({
            "nome":         nome.replace("\n", " "),
            "bairro":       bairro,
            "endereco":     endereco,
            "area":         str(row[COL_AREA] or ""),
            "quartos":      str(row[COL_QUARTOS] or ""),
            "diferenciais": str(row[COL_DIFER] or ""),
            "lazer":        str(row[COL_LAZER] or ""),
            "preco":        preco_final,
            "entrega":      formatar_entrega(row[COL_ENTREGA]),
            "imagem":       imagem_path,
            "lat":          lat,
            "lng":          lng,
        })

    # ══════════════════════════════════════════════════════════
    #  SAÍDA
    # ══════════════════════════════════════════════════════════

    if dry_run:
        # ── Modo Prévia ───────────────────────────────────────
        print("\n" + "─" * 55)
        print("📋 PRÉVIA (--dry-run): Nenhum arquivo foi alterado.\n")
        for i, emp in enumerate(empreendimentos):
            print(f"  {i+1}. {emp['nome']}")
            print(f"     📍 {emp['lat']:.5f}, {emp['lng']:.5f}")
            print(f"     💰 {emp['preco']} | 📅 {emp['entrega']}")
            print(f"     🖼️  {emp['imagem'] or '⚠️  SEM IMAGEM'}")
            print()
    else:
        # ── Backup do arquivo anterior ────────────────────────
        if os.path.exists(ARQUIVO_JS):
            backup_path = ARQUIVO_JS + ".bkp"
            try:
                import shutil
                shutil.copy2(ARQUIVO_JS, backup_path)
                print(f"\n💾 Backup criado: {backup_path}")
            except Exception as e:
                print(f"  ⚠️  Falha ao criar backup: {e}")

        # ── Gerar empreendimentos.js ──────────────────────────
        with open(ARQUIVO_JS, "w", encoding="utf-8") as f:
            f.write("var EMPREENDIMENTOS = ")
            json.dump(empreendimentos, f, ensure_ascii=False, indent=2)
            f.write(";\n")

    # ── Relatório Final ───────────────────────────────────────
    print("\n" + "=" * 55)
    print(f"✅ {len(empreendimentos)} empreendimentos processados!")
    if erros_geocod > 0:
        print(f"⚠️  {erros_geocod} endereço(s) com falha de geocodificação.")
    if imagens_faltantes:
        print(f"⚠️  {len(imagens_faltantes)} imagem(ns) não encontrada(s):")
        for img in imagens_faltantes:
            print(f"     ❌ {PASTA_IMAGENS}/{img}")
    if not dry_run:
        print(f"📄 Arquivo gerado: {ARQUIVO_JS}")
        print(f"🗺️  Mapa atualizado com {len(empreendimentos)} pins.")
    print("=" * 55)


# ── Variáveis de Ambiente (.env) ──────────────────────────────
# Se existir um arquivo .env na raiz com GOOGLE_MAPS_API_KEY,
# o script usará a API do Google Maps para geocodificação (mais precisa).
# Caso contrário, usará o Nominatim (gratuito, OpenStreetMap).
def carregar_env():
    """Carrega variáveis do arquivo .env (chave=valor)."""
    env = {}
    if os.path.exists(".env"):
        with open(".env", "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if "=" in line and not line.startswith("#") and not line.startswith("//"):
                    k, v = line.strip().split("=", 1)
                    # Remove possíveis aspas simples ou duplas do valor
                    clean_val = v.strip().strip("'").strip('"')
                    env[k.strip()] = clean_val
    if not env.get("GOOGLE_MAPS_API_KEY"):
        print("  ℹ️  GOOGLE_MAPS_API_KEY não encontrada. Usando Nominatim como padrão.")
    return env


ENV = carregar_env()
GOOGLE_KEY = ENV.get("GOOGLE_MAPS_API_KEY")

if __name__ == "__main__":
    main()
