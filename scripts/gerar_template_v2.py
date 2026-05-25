import openpyxl
from openpyxl.styles import Font, Fill, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, Reference

def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dossiê Inteligente MT"
    
    # 🎨 Brand Colors (Champagne & Dark)
    GOLD = "C5A059"
    DARK = "12141A"
    WHITE = "FFFFFF"
    
    # --- Estilização Geral ---
    def style_header(cell, text):
        cell.value = text
        cell.font = Font(name='Arial', size=16, bold=True, color=GOLD)
        cell.alignment = Alignment(horizontal='center')
        
    def style_label(cell, text):
        cell.value = text
        cell.font = Font(name='Arial', size=11, bold=True, color=WHITE)
        cell.fill = PatternFill(start_color="1F2129", end_color="1F2129", fill_type="solid")
        
    def style_input(cell):
        cell.fill = PatternFill(start_color="2A2D35", end_color="2A2D35", fill_type="solid")
        cell.font = Font(color=GOLD, bold=True)
        cell.alignment = Alignment(horizontal='right')

    # 1. Cabeçalho Principal
    ws.merge_cells('B2:H3')
    style_header(ws['B2'], "MT PARCEIROS | DOSSIÊ DE INTELIGÊNCIA PATRIMONIAL")
    
    # 2. SEÇÃO: DADOS DO CLIENTE (Inputs)
    ws['B5'] = "IDENTIFICAÇÃO E RENDA"
    ws['B5'].font = Font(bold=True, size=12, color=GOLD)
    
    labels = [
        ("C7", "NOME COMPLETO:"),
        ("C8", "RENDA MENSAL BRUTA:"),
        ("C9", "SALDO FGTS DISPONÍVEL:"),
        ("C10", "RECURSOS PRÓPRIOS (ENTRADA):"),
        ("C11", "DÍVIDAS/CONSIGNADOS:"),
        ("C12", "VÍNCULO EMPREGATÍCIO:"),
        ("C13", "POSSUI DEPENDENTES?"),
    ]
    
    for pos, text in labels:
        style_label(ws[pos], text)
        style_input(ws[pos.replace('C', 'E')])
        
    # 3. SEÇÃO: DOSSIÊ DNA (Property Details - Injected by JS)
    ws['G5'] = "ATRIBUTOS DO IMÓVEL (DNA)"
    ws['G5'].font = Font(bold=True, size=12, color=GOLD)
    
    dna_labels = [
        ("G7", "IMÓVEL SELECIONADO:"),
        ("G8", "AUDITORIA JURÍDICA:"),
        ("G9", "VISTORIA TÉCNICA:"),
        ("G10", "SCORE FINANCEIRO:"),
        ("G11", "POTENCIAL VALORIZAÇÃO:"),
    ]
    
    for pos, text in dna_labels:
        style_label(ws[pos], text)
        style_input(ws[pos.replace('G', 'H')])

    # 4. SEÇÃO: RESULTADO DA ANÁLISE (Dashboard)
    ws['B16'] = "RESULTADO DA ANÁLISE PROFISSIONAL"
    ws['B16'].font = Font(bold=True, size=14, color=GOLD)
    
    results = [
        ("C18", "PODER DE COMPRA ESTIMADO:"),
        ("C19", "SUBSÍDIO FEDERAL/ESTADUAL:"),
        ("C20", "VALOR MÁXIMO FINANCIÁVEL:"),
        ("C21", "TAXA DE JUROS ANUAL:"),
        ("C22", "PRAZO TOTAL (MESES):"),
    ]
    
    for pos, text in results:
        style_label(ws[pos], text)
        # Fórmulas de Placeholder (Serão alimentadas pelo JS ou Fórmulas Vivas)
        ws[pos.replace('C', 'E')].font = Font(bold=True, size=12, color=GOLD)
        ws[pos.replace('C', 'E')].number_format = '"R$ "#,##0'

    # 5. FLUXO DE PAGAMENTO (36 MESES)
    ws['G16'] = "FLUXO ESTIMADO DE PAGAMENTO (36 MESES)"
    ws['G16'].font = Font(bold=True, size=12, color=GOLD)
    
    fluxo = [
        ("G18", "PARCELA MÉDIA (OBRA/INCC):"),
        ("G19", "ATO / SINAL:"),
        ("G20", "PARCELA PÓS-CHAVES (SAC):"),
    ]
    
    for pos, text in fluxo:
        style_label(ws[pos], text)
        ws[pos.replace('G', 'H')].font = Font(bold=True, size=11, color="19E08F")
        ws[pos.replace('G', 'H')].number_format = '"R$ "#,##0'

    # 6. DICA DO CONSULTOR (MT Insight)
    ws.merge_cells('B25:H26')
    ws['B25'] = "DICA DO ESTRATEGISTA MT INSIGHT: Você pode aumentar seu potencial de financiamento em até 15% reduzindo dívidas ativas ou compondo renda com familiares."
    ws['B25'].font = Font(italic=True, size=11, color="CCCCCC")
    ws['B25'].alignment = Alignment(wrap_text=True, vertical='center')

    # Ajuste de Colunas
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 25
    
    # 🎨 Dark Background Emulation (Fill entire visible area)
    for row in range(1, 100):
        for col in range(1, 20):
            if not ws.cell(row=row, column=col).fill.patternType:
                ws.cell(row=row, column=col).fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")

    # Save
    wb.save("c:/Users/Marcos.PC_M1/Documents/site_mt/assets/docs/template_mt_parceiros_v2.xlsx")
    print("Sucesso: Template Excel v2 (Dashboard) gerado com sucesso.")

if __name__ == "__main__":
    create_template()
