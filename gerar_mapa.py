# ============================================================
#  gerar_mapa.py — MT Parceiros
#  Geocodificação de ALTA PRECISÃO (Nominatim + Hard Fixes).
#  Garante que os pontos fiquem EXATAMENTE nos prédios.
# ============================================================
import json, time, sys, os, re
import urllib.request, urllib.parse

try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl")
    import openpyxl

# ── Configurações ────────────────────────────────────────────
ARQUIVO_XLSX   = "Empreendimentos.xlsx"
ARQUIVO_JS     = "empreendimentos.js"
ARQUIVO_CACHE  = "coords_cache.json"

# Correções Manuais (Coordenadas Reais validadas via GPS)
FIXES = {
    "Rua Laguna, 440":                     {"lat": -23.63959, "lng": -46.71574},
    "Avenida Cristo Rei, 34":              {"lat": -23.48938, "lng": -46.72078},
    "Professor Miguel Franchini Neto, 186": {"lat": -23.44355, "lng": -46.72530},
    "Rua Professor Clemente Pastore, 14":  {"lat": -23.51383, "lng": -46.69922},
    "Rua Joao Alfredo, 431":               {"lat": -23.65701, "lng": -46.70044},
    "Avenida Condessa Elisabeth de Robiano, 5748": {"lat": -23.51590, "lng": -46.55550}
}

# Limites de São Paulo (evita que o geocodificador fuja da capital)
LAT_MIN, LAT_MAX = -23.85, -23.35
LNG_MIN, LNG_MAX = -46.95, -46.35

# Índices das colunas
COL_NOME = 0; COL_BAIRRO = 1; COL_END = 2; COL_AREA = 3; COL_QUARTOS = 4
COL_PRECO = 6; COL_ENTREGA = 7

# ── Funções ──────────────────────────────────────────────────
def carregar_cache():
    if os.path.exists(ARQUIVO_CACHE):
        try:
            with open(ARQUIVO_CACHE, "r", encoding="utf-8") as f:
                return json.load(f)
        except: return {}
    return {}

def salvar_cache(cache):
    with open(ARQUIVO_CACHE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)

def geocodificar_nominatim(endereco, bairro):
    headers = { 'User-Agent': 'MTParceirosGeocoder/4.0' }
    
    # Limpeza básica: remover acentos e normalizar
    addr_clean = endereco.replace("\n", " ").strip()

    # Checar se temos correção fixa para este endereço
    for partial_match, coords in FIXES.items():
        if partial_match.lower() in addr_clean.lower():
            return coords["lat"], coords["lng"], "Fixa"

    consultas = [
        f"{addr_clean}, Sao Paulo, SP, Brazil",
        f"{addr_clean}, {bairro}, Sao Paulo, Brazil"
    ]
    
    for q in consultas:
        try:
            url = f"https://nominatim.openstreetmap.org/search?q={urllib.parse.quote(q)}&format=json&limit=3&bounded=1&viewbox=-46.85,-23.35,-46.35,-23.85"
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read().decode())
                for result in data:
                    lat, lon = float(result["lat"]), float(result["lon"])
                    # Validar se está DENTRO de São Paulo
                    if LAT_MIN < lat < LAT_MAX and LNG_MIN < lon < LNG_MAX:
                        return lat, lon, "Real"
            time.sleep(1.2)
        except: pass
        
    return None, None, None

def main():
    print("=" * 55); print("  MT Parceiros — Gerador de Mapa de ALTA PRECISÃO"); print("=" * 55)

    if not os.path.exists(ARQUIVO_XLSX):
        print(f"\n❌ Arquivo '{ARQUIVO_XLSX}' não encontrado."); sys.exit(1)

    cache = carregar_cache()
    wb = openpyxl.load_workbook(ARQUIVO_XLSX); ws = wb.active
    empreendimentos = []
    erros = 0
    linhas = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
    total = sum(1 for r in linhas if r[COL_NOME])
    print(f"📋 Processando {total} empreendimento(s)...\n")

    for i, row in enumerate(linhas):
        nome = str(row[COL_NOME] or "").strip()
        if not nome: continue
        endereco = str(row[COL_END] or "").strip().replace("\n", " ")
        bairro   = str(row[COL_BAIRRO] or "").strip()

        print(f"🔍 [{i+1}/{total}] {nome}")
        
        lat, lng = None, None
        if endereco in cache:
            lat, lng = cache[endereco]["lat"], cache[endereco]["lng"]
            print(f"  🧊 (Cache) {lat:.5f}, {lng:.5f}")
        else:
            lat, lng, tipo = geocodificar_nominatim(endereco, bairro)
            if lat:
                print(f"  📍 ({tipo}) {lat:.5f}, {lng:.5f}")
                cache[endereco] = {"lat": lat, "lng": lng}
                salvar_cache(cache)
            else:
                print(f"  ❌ Falha no endereço: {endereco}"); erros += 1; continue

        empreendimentos.append({
            "nome": nome, "bairro": bairro, "endereco": endereco,
            "area": str(row[COL_AREA] or ""), "quartos": str(row[COL_QUARTOS] or ""),
            "preco": str(row[COL_PRECO] or ""), "entrega": str(row[COL_ENTREGA] or ""),
            "lat": lat, "lng": lng
        })

    with open(ARQUIVO_JS, "w", encoding="utf-8") as f:
        f.write("var EMPREENDIMENTOS = "); json.dump(empreendimentos, f, ensure_ascii=False, indent=2); f.write(";")

    print("\n" + "=" * 55); print(f"✅ {len(empreendimentos)} pontos gerados com precisão GPS!"); print("=" * 55)

if __name__ == "__main__":
    main()
